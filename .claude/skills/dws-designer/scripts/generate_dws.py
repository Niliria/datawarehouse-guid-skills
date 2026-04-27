#!/usr/bin/env python3
"""Generate DWS design outputs from DWD metadata and optional bus matrix files."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


OUTPUT_DOC_HEADER = [
    "表名称",
    "表备注",
    "字段名称",
    "字段类型",
    "字段备注",
    "来源表",
    "来源字段",
    "加工逻辑说明",
]


def read_csv_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore").replace("\r\n", "\n")


def parse_dwd_list(path: Path) -> list[dict]:
    text = read_csv_text(path)
    sections = re.split(r"\n\s*\n", text.strip())
    models: list[dict] = []

    for section in sections:
        rows = list(csv.reader(section.splitlines()))
        if not rows or len(rows[0]) < 2 or rows[0][0].strip() != "模型名":
            continue

        model_name = rows[0][1].strip()
        meta: dict[str, str] = {"模型名": model_name}
        fields: list[dict[str, str]] = []
        header: list[str] | None = None

        for row in rows[1:]:
            if not row or not any(cell.strip() for cell in row):
                continue
            key = row[0].strip()
            if key == "字段":
                header = [cell.strip() for cell in row]
                continue
            if header is None:
                if len(row) >= 2 and key:
                    meta[key] = row[1].strip()
                continue
            row = [cell.strip() for cell in row] + [""] * max(0, len(header) - len(row))
            field = {header[i]: row[i] for i in range(len(header))}
            if field.get("字段名"):
                fields.append(field)

        models.append({"meta": meta, "fields": fields})

    return models


def read_bus_matrix(path: Path) -> dict[str, list[list[str | None]]]:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to read XLSX files. Install with: pip install openpyxl"
        )

    workbook = load_workbook(path, data_only=True)
    matrix: dict[str, list[list[str | None]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        matrix[sheet_name] = [
            [cell.value for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    return matrix


def normalize_identifier(value: str) -> str:
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "_", value)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned.lower()


def map_field_type(field_type: str) -> str:
    ft = field_type.lower()
    if "decimal" in ft:
        return re.sub(r"varchar\(.*\)", "string", field_type, flags=re.IGNORECASE).upper()
    if "varchar" in ft or "string" in ft:
        return "STRING"
    if "bigint" in ft:
        return "BIGINT"
    if "int" in ft and "bigint" not in ft:
        return "INT"
    if "timestamp" in ft or "datetime" in ft:
        return "TIMESTAMP"
    return field_type.upper() if field_type else "STRING"


def build_dws_table_name(model_name: str) -> str:
    table_id = model_name
    if table_id.startswith("dwd_"):
        table_id = table_id[4:]
    if table_id.endswith("_di"):
        table_id = table_id[: -len("_di")]
    safe = normalize_identifier(table_id)
    return f"dws_{safe}_day_df"


def build_sql_file_name(table_name: str) -> str:
    safe = normalize_identifier(table_name)
    return f"{safe}.sql"


def build_doc_rows(model: dict) -> tuple[str, list[list[str]]]:
    model_name = model["meta"].get("模型名", "")
    table_name = build_dws_table_name(model_name)
    table_comment = f"DWS 汇总层-原子指标表 ({model_name})"
    fields = model["fields"]
    rows: list[list[str]] = []

    dims = [f for f in fields if f.get("维度", "").strip().upper() == "Y"]
    metrics = [f for f in fields if f.get("度量", "").strip().upper() == "Y"]

    if not any(f.get("字段名") == "pt" for f in fields):
        raise ValueError(f"模型 {model_name} 未包含 pt 分区字段，无法生成 DWS 表")

    rows.append(
        [
            table_name,
            table_comment,
            "dt",
            "STRING",
            "统计日期分区，来源 DWD.pt",
            model_name,
            "pt",
            "按日期分区输出 DWS 数据",
        ]
    )

    for dim in dims:
        name = dim["字段名"]
        rows.append(
            [
                table_name,
                table_comment,
                name,
                map_field_type(dim.get("字段类型", "STRING")),
                dim.get("字段说明", ""),
                model_name,
                name,
                "按主粒度维度保留外键字段",
            ]
        )

    for metric in metrics:
        name = metric["字段名"]
        rows.append(
            [
                table_name,
                table_comment,
                name,
                map_field_type(metric.get("字段类型", "STRING")),
                metric.get("字段说明", ""),
                model_name,
                name,
                "原子指标聚合(SUM)",
            ]
        )

    return table_name, rows


def build_ddl(model: dict) -> str:
    model_name = model["meta"].get("模型名", "")
    table_name = build_dws_table_name(model_name)
    fields = model["fields"]
    dims = [f for f in fields if f.get("维度", "").strip().upper() == "Y"]
    metrics = [f for f in fields if f.get("度量", "").strip().upper() == "Y"]

    columns: list[str] = []
    for dim in dims:
        name = dim["字段名"]
        columns.append(
            f"    {name} {map_field_type(dim.get('字段类型', 'STRING'))} COMMENT '{dim.get('字段说明', '')}'"
        )

    for metric in metrics:
        name = metric["字段名"]
        columns.append(
            f"    {name} {map_field_type(metric.get('字段类型', 'STRING'))} COMMENT '{metric.get('字段说明', '')}'"
        )

    columns_sql = ",\n".join(columns)
    return (
        f"CREATE TABLE IF NOT EXISTS {table_name} (\n"
        f"{columns_sql}\n"
        f")\n"
        f"COMMENT 'DWS 汇总层表，存储原子指标，不包含复合/派生指标'\n"
        f"PARTITIONED BY (dt STRING COMMENT '统计日期(YYYY-MM-DD)')\n"
        f"STORED AS ORC\n"
        f"TBLPROPERTIES (\n"
        f"    'orc.compress'='SNAPPY',\n"
        f"    'orc.stripe.size'='67108864'\n"
        f");\n"
    )


def build_etl(model: dict) -> str:
    model_name = model["meta"].get("模型名", "")
    table_name = build_dws_table_name(model_name)
    source_table = model_name
    fields = model["fields"]
    dims = [f["字段名"] for f in fields if f.get("维度", "").strip().upper() == "Y"]
    metrics = [f["字段名"] for f in fields if f.get("度量", "").strip().upper() == "Y"]

    group_cols = dims + ["pt"]
    select_columns = [f"    {col}" for col in dims]
    select_columns += [f"    SUM({col}) AS {col}" for col in metrics]
    select_columns.append("    pt AS dt")

    return (
        f"INSERT OVERWRITE TABLE {table_name} PARTITION(dt)\n"
        f"SELECT\n"
        + ",\n".join(select_columns)
        + f"\nFROM {source_table}\n"
        f"WHERE is_valid = 1\n"
        f"GROUP BY {', '.join(group_cols)};\n"
    )


def write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate DWS outputs from DWD metadata.")
    parser.add_argument(
        "--input-dwd-csv",
        type=Path,
        default=Path("output/cdm_modeling_skill/docs/dwd_list.csv"),
        help="DWD 元数据清单 CSV 文件",
    )
    parser.add_argument(
        "--input-dim-csv",
        type=Path,
        default=Path("output/cdm_modeling_skill/docs/dim_list.csv"),
        help="DIM 元数据清单 CSV 文件，仅用于校验和扩展",
    )
    parser.add_argument(
        "--input-bus-matrix-xlsx",
        type=Path,
        default=Path("output/dwm-bus-matrix/dwm_bus_matrix.xlsx"),
        help="可选的总线矩阵 XLSX 文件，用于主题校验",
    )
    parser.add_argument(
        "--output-doc-csv",
        type=Path,
        default=Path("output/dws-designer/docs/dws_list.csv"),
        help="生成的 DWS 设计文档 CSV",
    )
    parser.add_argument(
        "--output-ddl-dir",
        type=Path,
        default=Path("output/dws-designer/ddl/dws"),
        help="生成的 DDL 目录",
    )
    parser.add_argument(
        "--output-etl-dir",
        type=Path,
        default=Path("output/dws-designer/etl/dws"),
        help="生成的 ETL 目录",
    )
    args = parser.parse_args()

    if not args.input_dwd_csv.exists():
        raise FileNotFoundError(f"缺少 DWD 元数据文件: {args.input_dwd_csv}")

    dwd_models = parse_dwd_list(args.input_dwd_csv)
    if not dwd_models:
        raise ValueError(f"从 {args.input_dwd_csv} 未解析到任何模型定义")

    if args.input_bus_matrix_xlsx.exists():
        try:
            matrix = read_bus_matrix(args.input_bus_matrix_xlsx)
            print(f"读取总线矩阵 XLSX: {args.input_bus_matrix_xlsx}，包含表: {list(matrix.keys())}")
        except Exception as exc:
            print(f"警告：读取总线矩阵失败: {exc}")
    else:
        print(f"未找到总线矩阵 XLSX，跳过该文件: {args.input_bus_matrix_xlsx}")

    if args.input_dim_csv.exists():
        print(f"DIM 元数据文件存在: {args.input_dim_csv}")
    else:
        print(f"WARN: DIM 元数据文件未找到: {args.input_dim_csv}")

    all_doc_rows: list[list[str]] = []

    args.output_ddl_dir.mkdir(parents=True, exist_ok=True)
    args.output_etl_dir.mkdir(parents=True, exist_ok=True)

    for model in dwd_models:
        table_name, doc_rows = build_doc_rows(model)
        ddl = build_ddl(model)
        etl = build_etl(model)

        ddl_path = args.output_ddl_dir / build_sql_file_name(table_name)
        etl_path = args.output_etl_dir / f"load_{build_sql_file_name(table_name)}"

        ddl_path.write_text(ddl, encoding="utf-8")
        etl_path.write_text(etl, encoding="utf-8")
        all_doc_rows.extend(doc_rows)

        print(f"生成 DDL: {ddl_path}")
        print(f"生成 ETL: {etl_path}")

    write_csv(args.output_doc_csv, OUTPUT_DOC_HEADER, all_doc_rows)
    print(f"生成 DWS 文档: {args.output_doc_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
