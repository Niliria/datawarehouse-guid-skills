#!/usr/bin/env python3
"""
Bus matrix Excel generator for dwm-bus-matrix skill.

Reads step3/step4 CSV outputs and generates a cross-tab xlsx file.

Usage:
    python write_bus_matrix.py \
        --table-profile output/dwm-bus-matrix/step3/dwm_s3_table_profile.csv \
        --subject-area  output/dwm-bus-matrix/step3/dwm_s3_subject_area.csv \
        --fact-dim-ref  output/dwm-bus-matrix/step4/dwm_s4_fact_dim_ref.csv \
        --dim-registry  output/dwm-bus-matrix/step4/dwm_s4_dim_registry.csv \
        --output        output/dwm-bus-matrix/dwm_bus_matrix.xlsx \
        --version       v1.0

Requires: openpyxl (pip install openpyxl)
"""

import csv
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def read_csv(filepath: str) -> list[dict]:
    """Read CSV with BOM handling."""
    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def build_matrix(
    table_profile: list[dict],
    subject_area: list[dict],
    fact_dim_ref: list[dict],
    dim_registry: list[dict],
) -> tuple[list[dict], list[dict], dict[tuple[str, str], str]]:
    """
    Build bus matrix data structures.

    Returns:
        - rows: list of fact business processes (sorted by subject area)
        - columns: list of conformed dimensions
        - cells: dict of (bp_standard_name, dimension_key) -> marker
    """
    # Build subject area lookup
    sa_lookup = {sa["subject_area_code"]: sa for sa in subject_area}

    # Build rows: fact tables with their business processes
    rows = []
    for tp in table_profile:
        if tp.get("table_role") != "fact":
            continue
        sa_code = tp.get("subject_area_code", "")
        sa_name = sa_lookup.get(sa_code, {}).get("subject_area_name_cn", sa_code)
        rows.append({
            "subject_area_code": sa_code,
            "subject_area_name": sa_name,
            "bp_standard_name": tp.get("bp_standard_name", ""),
            "fact_type": tp.get("fact_type", ""),
            "ods_table_name": tp.get("ods_table_name", ""),
        })

    # Sort by subject area, then by bp_standard_name
    rows.sort(key=lambda r: (r["subject_area_code"], r["bp_standard_name"]))

    # Build columns: conformed dimensions from dim_registry
    columns = sorted(dim_registry, key=lambda d: d.get("dimension_key", ""))

    # Build dimension_key lookup from ref_table
    ref_to_dim = {}
    for dim in dim_registry:
        src_table = dim.get("source_dimension_table", "")
        if src_table:
            ref_to_dim[src_table] = dim.get("dimension_key", "")

    # Build cells: which fact connects to which dimension
    cells: dict[tuple[str, str], str] = {}
    for ref in fact_dim_ref:
        if ref.get("dimension_type") != "外键":
            continue
        fact_table = ref.get("fact_table", "")
        ref_table = ref.get("ref_table", "")
        dim_key = ref_to_dim.get(ref_table, "")
        if not dim_key:
            continue

        # Find bp_standard_name for this fact_table
        bp_name = ref.get("bp_standard_name", "")
        if bp_name and dim_key:
            cells[(bp_name, dim_key)] = "✓"

    return rows, columns, cells


def write_xlsx(
    rows: list[dict],
    columns: list[dict],
    cells: dict[tuple[str, str], str],
    output_path: str,
    version: str = "v1.0",
    status: str = "draft",
):
    """Generate the bus matrix Excel file."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"总线矩阵 {version}"

    # -- Styles --
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    check_font = Font(size=12, bold=True, color="2E7D32")
    dash_font = Font(size=12, color="BDBDBD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # -- Header row --
    fixed_headers = ["主题域", "业务过程", "事实表类型"]
    dim_headers = [d.get("dimension_name", d.get("dimension_key", "")) for d in columns]
    all_headers = fixed_headers + dim_headers

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # -- Data rows --
    data_start_row = 2

    for row_idx, row_data in enumerate(rows, data_start_row):
        sa_code = row_data["subject_area_code"]
        sa_name = row_data["subject_area_name"]
        bp_name = row_data["bp_standard_name"]
        fact_type = row_data["fact_type"]

        # Column A: subject area (every row)
        cell_a = ws.cell(row=row_idx, column=1, value=f"{sa_name}({sa_code})")
        cell_a.alignment = Alignment(vertical="center")
        cell_a.border = thin_border

        # Column B: business process
        cell_b = ws.cell(row=row_idx, column=2, value=bp_name)
        cell_b.border = thin_border

        # Column C: fact type
        cell_c = ws.cell(row=row_idx, column=3, value=fact_type)
        cell_c.border = thin_border

        # Dimension columns
        for dim_idx, dim in enumerate(columns):
            dim_key = dim.get("dimension_key", "")
            marker = cells.get((bp_name, dim_key), "-")
            cell = ws.cell(row=row_idx, column=4 + dim_idx, value=marker)
            cell.alignment = center_align
            cell.border = thin_border
            if marker == "✓":
                cell.font = check_font
            else:
                cell.font = dash_font

    last_data_row = data_start_row + len(rows) - 1

    # -- Metadata footer --
    meta_row = last_data_row + 2 if rows else 3
    meta_items = [
        ("version", version),
        ("status", status),
        ("updated_by", "dwm-bus-matrix-skill"),
        ("updated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    meta_font = Font(italic=True, color="888888", size=9)
    for i, (key, val) in enumerate(meta_items):
        ws.cell(row=meta_row, column=1 + i * 2, value=key).font = meta_font
        ws.cell(row=meta_row, column=2 + i * 2, value=val).font = meta_font

    # -- Column widths --
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    for i in range(len(columns)):
        col_letter = chr(ord("D") + i) if i < 23 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 14

    wb.save(output_path)
    print(f"Bus matrix written -> {output_path} ({len(rows)} processes × {len(columns)} dimensions)")


def main():
    parser = argparse.ArgumentParser(description="Generate bus matrix Excel from step3/step4 CSVs")
    parser.add_argument("--table-profile", required=True, help="Path to dwm_s3_table_profile.csv")
    parser.add_argument("--subject-area", required=True, help="Path to dwm_s3_subject_area.csv")
    parser.add_argument("--fact-dim-ref", required=True, help="Path to dwm_s4_fact_dim_ref.csv")
    parser.add_argument("--dim-registry", required=True, help="Path to dwm_s4_dim_registry.csv")
    parser.add_argument("--output", "-o", default="output/dwm-bus-matrix/dwm_bus_matrix.xlsx",
                        help="Output xlsx path (default: output/dwm-bus-matrix/dwm_bus_matrix.xlsx)")
    parser.add_argument("--version", "-v", default="v1.0", help="Matrix version (default: v1.0)")
    parser.add_argument("--status", "-s", default="draft", help="Matrix status (default: draft)")
    args = parser.parse_args()

    table_profile = read_csv(args.table_profile)
    subject_area = read_csv(args.subject_area)
    fact_dim_ref = read_csv(args.fact_dim_ref)
    dim_registry = read_csv(args.dim_registry)

    rows, columns, cells = build_matrix(table_profile, subject_area, fact_dim_ref, dim_registry)
    write_xlsx(rows, columns, cells, args.output, args.version, args.status)


if __name__ == "__main__":
    main()
