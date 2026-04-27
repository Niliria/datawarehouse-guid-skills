
import os
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict


def load_metadata_from_excel(file_path):
    """
    从Excel文件加载元数据
    
    Args:
        file_path (str): Excel文件路径
        
    Returns:
        tuple: (表信息字典, 字段信息列表)
    """
    tables = {}
    fields = []
    
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        return tables, fields
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        data = {
            'data_source': ws.cell(row=row, column=1).value,
            'db_name': ws.cell(row=row, column=2).value,
            'table_name': ws.cell(row=row, column=3).value,
            'table_chinese_name': ws.cell(row=row, column=4).value,
            'field': ws.cell(row=row, column=5).value,
            'comment': ws.cell(row=row, column=6).value,
            'comment_fill': ws.cell(row=row, column=7).value,
            'data_type': ws.cell(row=row, column=8).value,
            'null': ws.cell(row=row, column=9).value,
            'default': ws.cell(row=row, column=10).value,
            'is_primary': ws.cell(row=row, column=11).value,
            'is_foreign': ws.cell(row=row, column=12).value,
            'foreign_key_ref': ws.cell(row=row, column=13).value,
            'fill_rate': ws.cell(row=row, column=14).value
        }
        
        fields.append(data)
        
        table_name = data['table_name']
        if table_name not in tables:
            tables[table_name] = {
                'table_name': table_name,
                'table_chinese_name': data['table_chinese_name'],
                'data_source': data['data_source'],
                'db_name': data['db_name'],
                'fields': []
            }
        tables[table_name]['fields'].append(data)
    
    wb.close()
    return tables, fields


def deduplicate_relations(relations):
    """
    去除重复的关系
    
    Args:
        relations (list): 关系列表
        
    Returns:
        list: 去重后的关系列表
    """
    seen = set()
    unique_relations = []
    
    for rel in relations:
        # 创建唯一键
        key = (
            rel['source_table'],
            rel['source_field'],
            rel['target_table'],
            rel['target_field']
        )
        
        if key not in seen:
            seen.add(key)
            unique_relations.append(rel)
    
    return unique_relations


def infer_foreign_key_relations(tables, fields):
    """
    通过字段名模式推断外键关系
    
    Args:
        tables (dict): 表信息字典
        fields (list): 字段信息列表
        
    Returns:
        list: 推断的外键关系列表
    """
    relations = []
    
    # 收集所有表的主键字段
    primary_keys = {}
    for table_name in tables:
        for field in tables[table_name]['fields']:
            if field['is_primary'] == 'YES':
                if table_name not in primary_keys:
                    primary_keys[table_name] = []
                primary_keys[table_name].append(field['field'])
    
    # 推断外键关系
    for field in fields:
        # 跳过主键
        if field['is_primary'] == 'YES':
            continue
        
        field_name = field['field']
        table_name = field['table_name']
        
        # 检查字段名是否以 _id 结尾
        if field_name.endswith('_id'):
            # 尝试匹配可能的源表
            base_name = field_name[:-3]  # 去掉 _id
            
            # 可能的源表名模式
            possible_table_names = [
                f"t_{base_name}",
                f"t_{base_name}s",
                f"t_{base_name[:-1]}" if base_name.endswith('s') else f"t_{base_name}s"
            ]
            
            # 尝试直接去掉前缀匹配
            if table_name.startswith('t_'):
                possible_table_names.append(f"t_{base_name}")
            
            # 去重可能的表名
            possible_table_names = list(dict.fromkeys(possible_table_names))
            
            # 查找可能的源表
            for possible_table in possible_table_names:
                if possible_table in primary_keys:
                    for pk_field in primary_keys[possible_table]:
                        # 如果主键字段名与当前字段名相似，或者主键字段是 id
                        if pk_field == field_name or pk_field == 'id' or pk_field == base_name + '_id':
                            relations.append({
                                'source_table': possible_table,
                                'source_field': pk_field,
                                'target_table': table_name,
                                'target_field': field_name,
                                'source_table_chinese': tables.get(possible_table, {}).get('table_chinese_name', ''),
                                'target_table_chinese': field['table_chinese_name'],
                                'relation_type': '推断外键'
                            })
        
        # 检查是否是常见的外键字段（如 user_id, enterprise_id 等）
        common_fk_patterns = [
            ('user_id', 't_user', 'user_id'),
            ('user_id', 't_user', 'id'),
            ('enterprise_id', 't_enterprise', 'enterprise_id'),
            ('enterprise_id', 't_enterprise', 'id'),
            ('role_id', 't_role', 'role_id'),
            ('role_id', 't_role', 'id'),
            ('menu_id', 't_menu', 'menu_id'),
            ('menu_id', 't_menu', 'id'),
            ('building_id', 't_building', 'building_number'),
            ('building_id', 't_building', 'id'),
            ('sg_user_id', 't_sg_user', 'user_id'),
            ('sg_user_id', 't_sg_user', 'id'),
        ]
        
        for pattern in common_fk_patterns:
            pattern_field, pattern_table, pattern_pk = pattern
            if field_name == pattern_field and pattern_table in tables:
                # 检查源表是否有这个主键
                has_pk = False
                for f in tables[pattern_table]['fields']:
                    if f['field'] == pattern_pk:
                        has_pk = True
                        break
                if has_pk:
                    relations.append({
                        'source_table': pattern_table,
                        'source_field': pattern_pk,
                        'target_table': table_name,
                        'target_field': field_name,
                        'source_table_chinese': tables.get(pattern_table, {}).get('table_chinese_name', ''),
                        'target_table_chinese': field['table_chinese_name'],
                        'relation_type': '推断外键'
                    })
    
    # 去重
    relations = deduplicate_relations(relations)
    return relations


def build_lineage_relationships(tables, fields):
    """
    构建血缘关系
    
    Args:
        tables (dict): 表信息字典
        fields (list): 字段信息列表
        
    Returns:
        dict: 血缘关系数据
    """
    # 首先尝试从原始数据获取外键关系
    foreign_key_relations = []
    
    for field in fields:
        if field['is_foreign'] == 'YES' and field['foreign_key_ref']:
            fk_ref = field['foreign_key_ref']
            if '.' in fk_ref:
                ref_table, ref_field = fk_ref.split('.', 1)
                foreign_key_relations.append({
                    'source_table': ref_table,
                    'source_field': ref_field,
                    'target_table': field['table_name'],
                    'target_field': field['field'],
                    'source_table_chinese': tables.get(ref_table, {}).get('table_chinese_name', ''),
                    'target_table_chinese': field['table_chinese_name'],
                    'relation_type': '外键引用'
                })
    
    # 如果没有找到外键关系，尝试通过字段名推断
    if not foreign_key_relations:
        print("未找到明确的外键引用，尝试通过字段名模式推断...")
        foreign_key_relations = infer_foreign_key_relations(tables, fields)
    
    # 表依赖关系
    table_dependencies = defaultdict(list)
    for rel in foreign_key_relations:
        if rel['source_table'] not in table_dependencies[rel['target_table']]:
            table_dependencies[rel['target_table']].append(rel['source_table'])
    
    # 字段血缘链路
    field_lineage_chains = []
    
    for field in fields:
        # 查找直接来源字段
        direct_sources = []
        for fk_rel in foreign_key_relations:
            if fk_rel['target_table'] == field['table_name'] and fk_rel['target_field'] == field['field']:
                direct_sources.append({
                    'table': fk_rel['source_table'],
                    'field': fk_rel['source_field'],
                    'table_chinese': fk_rel['source_table_chinese']
                })
        
        if direct_sources:
            for source in direct_sources:
                field_lineage_chains.append({
                    'level': 1,
                    'source_table': source['table'],
                    'source_field': source['field'],
                    'source_table_chinese': source['table_chinese'],
                    'target_table': field['table_name'],
                    'target_field': field['field'],
                    'target_table_chinese': field['table_chinese_name'],
                    'lineage_path': f"{source['table']}.{source['field']} → {field['table_name']}.{field['field']}"
                })
    
    return {
        'foreign_key_relations': foreign_key_relations,
        'table_dependencies': table_dependencies,
        'field_lineage_chains': field_lineage_chains
    }


def create_excel_report(tables, fields, lineage_data, output_path):
    """
    创建Excel血缘分析报告
    
    Args:
        tables (dict): 表信息字典
        fields (list): 字段信息列表
        lineage_data (dict): 血缘关系数据
        output_path (str): 输出文件路径
    """
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. 表结构概览
    ws1 = wb.create_sheet("表结构概览")
    ws1.title = "表结构概览"
    
    headers1 = ["序号", "表名", "表中文名", "数据源类型", "业务数据库名称", "字段数量"]
    ws1.append(headers1)
    
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, table_name in enumerate(sorted(tables.keys()), 1):
        table = tables[table_name]
        ws1.append([
            idx,
            table['table_name'],
            table['table_chinese_name'],
            table['data_source'],
            table['db_name'],
            len(table['fields'])
        ])
    
    # 2. 字段详情
    ws2 = wb.create_sheet("字段详情")
    ws2.title = "字段详情"
    
    headers2 = ["序号", "表名", "表中文名", "字段名", "字段注释", "字段注释填充", "数据类型", "是否为空", "默认值", "主键", "外键", "外键引用"]
    ws2.append(headers2)
    
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, field in enumerate(fields, 1):
        ws2.append([
            idx,
            field['table_name'],
            field['table_chinese_name'],
            field['field'],
            field['comment'],
            field['comment_fill'],
            field['data_type'],
            field['null'],
            field['default'],
            field['is_primary'],
            field['is_foreign'],
            field['foreign_key_ref']
        ])
    
    # 3. 外键引用关系
    ws3 = wb.create_sheet("外键引用关系")
    ws3.title = "外键引用关系"
    
    headers3 = ["序号", "源表", "源表中文名", "源字段", "目标表", "目标表中文名", "目标字段", "关系类型"]
    ws3.append(headers3)
    
    for col_idx, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, rel in enumerate(lineage_data['foreign_key_relations'], 1):
        ws3.append([
            idx,
            rel['source_table'],
            rel['source_table_chinese'],
            rel['source_field'],
            rel['target_table'],
            rel['target_table_chinese'],
            rel['target_field'],
            rel['relation_type']
        ])
    
    # 4. 表依赖关系
    ws4 = wb.create_sheet("表依赖关系")
    ws4.title = "表依赖关系"
    
    headers4 = ["序号", "目标表", "目标表中文名", "依赖表", "依赖表中文名", "依赖层级"]
    ws4.append(headers4)
    
    for col_idx, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    idx = 1
    for target_table in sorted(lineage_data['table_dependencies'].keys()):
        source_tables = lineage_data['table_dependencies'][target_table]
        for source_table in source_tables:
            ws4.append([
                idx,
                target_table,
                tables.get(target_table, {}).get('table_chinese_name', ''),
                source_table,
                tables.get(source_table, {}).get('table_chinese_name', ''),
                1
            ])
            idx += 1
    
    # 5. 字段血缘链路
    ws5 = wb.create_sheet("字段血缘链路")
    ws5.title = "字段血缘链路"
    
    headers5 = ["序号", "层级", "源表", "源表中文名", "源字段", "目标表", "目标表中文名", "目标字段", "血缘路径"]
    ws5.append(headers5)
    
    for col_idx, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, chain in enumerate(lineage_data['field_lineage_chains'], 1):
        ws5.append([
            idx,
            chain['level'],
            chain['source_table'],
            chain['source_table_chinese'],
            chain['source_field'],
            chain['target_table'],
            chain['target_table_chinese'],
            chain['target_field'],
            chain['lineage_path']
        ])
    
    # 6. 数据字典总览
    ws6 = wb.create_sheet("数据字典总览")
    ws6.title = "数据字典总览"
    
    headers6 = ["表名", "表中文名", "字段名", "字段注释", "数据类型", "是否为空", "是否主键", "是否外键"]
    ws6.append(headers6)
    
    for col_idx, header in enumerate(headers6, 1):
        cell = ws6.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for field in fields:
        is_foreign_key = "YES" if any(rel['target_table'] == field['table_name'] and rel['target_field'] == field['field'] 
                                        for rel in lineage_data['foreign_key_relations']) else "NO"
        ws6.append([
            field['table_name'],
            field['table_chinese_name'],
            field['field'],
            field['comment_fill'],
            field['data_type'],
            field['null'],
            field['is_primary'],
            is_foreign_key
        ])
    
    # 调整列宽
    column_widths_map = {
        "表结构概览": [8, 25, 25, 15, 20, 10],
        "字段详情": [8, 25, 25, 25, 30, 30, 20, 10, 15, 8, 8, 30],
        "外键引用关系": [8, 25, 25, 20, 25, 25, 20, 12],
        "表依赖关系": [8, 25, 25, 25, 25, 10],
        "字段血缘链路": [8, 8, 25, 25, 20, 25, 25, 20, 50],
        "数据字典总览": [25, 25, 25, 30, 20, 10, 10, 10]
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name in column_widths_map:
            widths = column_widths_map[sheet_name]
            for idx, width in enumerate(widths, 1):
                if idx <= ws.max_column:
                    ws.column_dimensions[chr(64 + idx)].width = width
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"✅ 元数据血缘分析报告已导出到: {output_path}")


def main():
    """
    主函数
    """
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../..'))
    input_file = os.path.join(project_root, 'Output', 'all_tables_metadata.xlsx')
    output_dir = os.path.join(project_root, 'Output')
    output_file = os.path.join(output_dir, 'metadata_field_lineage.xlsx')
    
    try:
        print("正在加载元数据...")
        tables, fields = load_metadata_from_excel(input_file)
        print(f"✅ 加载了 {len(tables)} 张表，{len(fields)} 个字段")
        
        print("正在构建血缘关系...")
        lineage_data = build_lineage_relationships(tables, fields)
        print(f"✅ 构建了 {len(lineage_data['foreign_key_relations'])} 个外键引用关系")
        print(f"✅ 构建了 {len(lineage_data['field_lineage_chains'])} 个字段血缘链路")
        
        print("正在生成Excel报告...")
        create_excel_report(tables, fields, lineage_data, output_file)
        
        print("\n🎉 元数据字段血缘解析完成!")
        
    except Exception as e:
        print(f"❌ 操作失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
