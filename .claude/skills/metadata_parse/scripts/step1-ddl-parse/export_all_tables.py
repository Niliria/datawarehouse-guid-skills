import yaml
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import re


def get_db_connection(config_name):
    """
    获取数据库连接
    
    Args:
        config_name (str): 配置名称，对应config.yaml中的数据库配置
        
    Returns:
        tuple: (数据库连接对象, 数据库配置字典)
    """
    # 构建配置文件路径（项目根目录）
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../..'))
    config_path = os.path.join(project_root, 'input/metadata_parse', 'config.yaml')
    # 读取配置文件
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    # 获取指定配置的数据库连接信息
    db_config = config['database'][config_name]

    # 建立数据库连接
    connection = pymysql.connect(
        host=db_config['host'],
        port=db_config['port'],
        user=db_config['user'],
        password=db_config['password'],
        database=db_config['dbname'],
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    return connection, db_config


def get_all_tables(connection):
    """
    获取数据库中所有表名
    
    Args:
        connection: 数据库连接对象
        
    Returns:
        list: 表名列表
    """
    with connection.cursor() as cursor:
        # 执行SQL查询获取所有表名
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()

    # 提取表名并返回列表
    return [list(table.values())[0] for table in tables]


def get_table_metadata(connection, table_name):
    """
    获取表的元数据
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        
    Returns:
        list: 字段元数据列表
    """
    with connection.cursor() as cursor:
        # 执行DESCRIBE语句获取表结构
        cursor.execute(f"DESCRIBE {table_name}")
        columns = cursor.fetchall()

    return columns


def get_column_comment(connection, table_name, column_name):
    """
    获取字段注释
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        column_name (str): 字段名
        
    Returns:
        str: 字段注释
    """
    with connection.cursor() as cursor:
        # 从INFORMATION_SCHEMA.COLUMNS查询字段注释
        cursor.execute("""
            SELECT COLUMN_COMMENT
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = %s
            AND COLUMN_NAME = %s
        """, (table_name, column_name))
        result = cursor.fetchone()

    # 返回字段注释，如果没有则返回空字符串
    return result['COLUMN_COMMENT'] if result and result['COLUMN_COMMENT'] else ""


def split_field_name(field_name):
    """
    拆解字段名
    
    Args:
        field_name (str): 字段名
        
    Returns:
        list: 拆解后的字段名部分列表
    """
    # 处理下划线分隔的字段名
    if '_' in field_name:
        parts = field_name.split('_')
        return [part.lower() for part in parts if part]

    # 处理驼峰命名和连续小写字母的字段名
    parts = []
    current_part = ''

    # 常见词汇列表，用于智能拆解字段名
    common_words = [
        'updated', 'created', 'deleted', 'enabled', 'disabled', 'visible', 'hidden',
        'shared', 'personal', 'global', 'local', 'default', 'custom', 'active', 'inactive',
        'classification', 'telephone', 'location', 'department', 'organization',
        'information', 'description', 'institution', 'administration', 'implementation',
        'configuration', 'documentation', 'communication', 'relationship', 'development',
        'management', 'performance', 'environment', 'infrastructure', 'authentication',
        'company', 'office', 'parent', 'depart', 'address', 'version', 'history', 'archive',
        'public', 'private',
        'update', 'create', 'delete', 'status', 'term', 'date', 'job', 'user',
        'name', 'code', 'type', 'time', 'id', 'number', 'email', 'phone',
        'url', 'path', 'size', 'cost', 'price', 'profit', 'value', 'count',
        'total', 'amount', 'unit', 'level', 'rank', 'order', 'product', 'service',
        'project', 'task', 'event', 'file', 'image', 'video', 'audio', 'role',
        'access', 'security', 'backup', 'restore', 'log', 'audit', 'alert', 'message',
        'comment', 'review', 'approval', 'reject', 'accept', 'submit', 'cancel', 'confirm',
        'verify', 'login', 'logout', 'register', 'reset', 'change', 'edit', 'view',
        'list', 'detail', 'report', 'dashboard', 'statistics', 'analysis', 'forecast',
        'trend', 'pattern', 'model', 'template', 'format', 'standard', 'rule', 'policy'
    ]

    # 去重并按长度降序排序，确保长词汇优先匹配
    unique_words = sorted(list(set(common_words)), key=len, reverse=True)

    i = 0
    n = len(field_name)

    while i < n:
        char = field_name[i]

        # 处理大写字母（驼峰命名）
        if char.isupper():
            if current_part:
                parts.append(current_part)
            current_part = char.lower()
            i += 1
        # 处理数字
        elif char.isdigit():
            if current_part and not current_part[-1].isdigit():
                parts.append(current_part)
                current_part = char
            else:
                current_part += char
            i += 1
        # 处理特殊字符
        elif not char.isalnum():
            if current_part:
                parts.append(current_part)
                current_part = ''
            i += 1
        # 处理小写字母
        else:
            matched = False
            # 尝试匹配最长的词汇
            for j in range(min(i + 20, n), i, -1):
                substring = field_name[i:j].lower()
                if substring in unique_words:
                    if current_part:
                        parts.append(current_part)
                        current_part = ''
                    parts.append(substring)
                    i = j
                    matched = True
                    break

            # 如果没有匹配到词汇，继续累积
            if not matched:
                current_part += char.lower()
                i += 1

    # 添加最后一个部分
    if current_part:
        parts.append(current_part)

    # 如果没有拆分成任何部分，返回原字段名的小写形式
    if not parts:
        parts = [field_name.lower()]

    return parts


def load_word_map():
    """
    从Excel参考文档加载词汇映射
    
    Returns:
        dict: 词汇映射字典
    """
    from openpyxl import load_workbook
    import os

    # 构建词汇映射文件路径（项目根目录）
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../..'))
    word_map_file = os.path.join(project_root, 'input/metadata_parse', 'word_map_reference.xlsx')

    word_map = {}

    try:
        # 加载Excel文件
        wb = load_workbook(word_map_file)
        ws = wb.active

        # 从第二行开始读取数据（第一行是表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # 确保英文词汇和中文翻译都不为空
                word_map[row[0].lower()] = row[1]

    except Exception as e:
        # 如果文件不存在或读取失败，使用默认词汇映射
        print(f"警告: 无法加载词汇映射文件，使用默认映射: {e}")
        word_map = {
            'id': 'ID',
            'name': '名称',
            'code': '代码'
        }

    return word_map


def translate_field_name(field_name):
    """
    翻译字段名为中文
    
    Args:
        field_name (str): 字段名
        
    Returns:
        str: 翻译后的中文名称
    """
    # 加载词汇映射
    word_map = load_word_map()

    # 拆解字段名
    parts = split_field_name(field_name)

    # 翻译每个部分
    translated_parts = []
    for part in parts:
        if part in word_map:
            translated_parts.append(word_map[part])
        else:
            # 对于未识别的部分，保留原词
            translated_parts.append(part)

    # 组合翻译结果
    return ''.join(translated_parts)


def get_table_chinese_name(connection, table_name):
    """
    从DDL语句中获取表的中文名
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        
    Returns:
        str: 表的中文名称
    """
    with connection.cursor() as cursor:
        try:
            # 执行SHOW CREATE TABLE获取DDL语句
            cursor.execute(f"SHOW CREATE TABLE {table_name}")
            result = cursor.fetchone()
            if result:
                ddl = result['Create Table']
                # 从DDL语句中提取表注释
                match = re.search(r"COMMENT='([^']+)'", ddl, re.IGNORECASE)
                if match:
                    comment = match.group(1).strip()
                    if comment:
                        return comment
            return ""
        except Exception as e:
            # 处理异常，返回空字符串
            return ""


def get_column_fill_rate(connection, table_name, column_name):
    """
    计算字段空值率
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        column_name (str): 字段名
        
    Returns:
        str: 空值率百分比
    """
    with connection.cursor() as cursor:
        try:
            # 获取总行数
            cursor.execute(f"SELECT COUNT(*) as total FROM {table_name}")
            total_result = cursor.fetchone()
            total = total_result['total']

            if total == 0:
                return "无数据"

            # 获取非空值行数
            cursor.execute(f"SELECT COUNT(*) as non_null_count FROM {table_name} WHERE {column_name} IS NOT NULL AND TRIM({column_name}) != ''")
            non_null_result = cursor.fetchone()
            non_null_count = non_null_result['non_null_count']

            # 计算空值数和空值率
            null_count = total - non_null_count
            null_rate = (null_count / total) * 100
            return f"{null_rate:.2f}%"
        except Exception as e:
            # 处理异常，返回计算失败
            return "计算失败"

    return "N/A"


def get_foreign_key_reference(connection, table_name, column_name):
    """
    获取外键引用信息
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        column_name (str): 字段名
        
    Returns:
        str: 外键引用信息，格式为"表名.字段名"
    """
    with connection.cursor() as cursor:
        try:
            # 查询外键引用信息
            cursor.execute("""
                SELECT
                    referenced_table_name,
                    referenced_column_name
                FROM
                    INFORMATION_SCHEMA.KEY_COLUMN_USAGE
                WHERE
                    table_schema = DATABASE()
                    AND table_name = %s
                    AND column_name = %s
                    AND referenced_table_name IS NOT NULL
                    AND referenced_column_name IS NOT NULL
            """, (table_name, column_name))

            result = cursor.fetchone()

            if result:
                return f"{result['referenced_table_name']}.{result['referenced_column_name']}"
            else:
                return ""
        except Exception as e:
            # 处理异常，返回空字符串
            return ""

    return ""


def get_table_row_count(connection, table_name):
    """
    获取表的总行数
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        
    Returns:
        int: 表的总行数
    """
    with connection.cursor() as cursor:
        try:
            cursor.execute(f"SELECT COUNT(*) as total FROM {table_name}")
            result = cursor.fetchone()
            return result['total'] if result else 0
        except Exception as e:
            print(f"警告: 获取 {table_name} 行数失败: {e}")
            return 0


def get_column_stats(connection, table_name, column_name):
    """
    获取字段的统计信息（总行数、非空值数、唯一值数等
    
    Args:
        connection: 数据库连接对象
        table_name (str): 表名
        column_name (str): 字段名
        
    Returns:
        dict: 字段统计信息字典
    """
    stats = {
        'total_cnt': 0,
        'non_null_cnt': 0,
        'null_cnt': 0,
        'null_rate': 0.0,
        'distinct_cnt': 0,
        'distinct_rate': 0.0
    }
    
    with connection.cursor() as cursor:
        try:
            # 获取总行数
            cursor.execute(f"SELECT COUNT(*) as total FROM {table_name}")
            total_result = cursor.fetchone()
            stats['total_cnt'] = total_result['total']
            
            if stats['total_cnt'] == 0:
                return stats
            
            # 获取非空值数
            cursor.execute(f"SELECT COUNT(*) as non_null_count FROM {table_name} WHERE {column_name} IS NOT NULL")
            non_null_result = cursor.fetchone()
            stats['non_null_cnt'] = non_null_result['non_null_count']
            stats['null_cnt'] = stats['total_cnt'] - stats['non_null_cnt']
            stats['null_rate'] = stats['null_cnt'] / stats['total_cnt'] * 100
            
            # 获取唯一值数
            cursor.execute(f"SELECT COUNT(DISTINCT {column_name}) as distinct_count FROM {table_name}")
            distinct_result = cursor.fetchone()
            stats['distinct_cnt'] = distinct_result['distinct_count']
            stats['distinct_rate'] = stats['distinct_cnt'] / stats['total_cnt'] * 100
            
        except Exception as e:
            print(f"警告: 获取 {table_name}.{column_name} 统计信息获取失败: {e}")
    
    return stats


def determine_field_role(field_name, data_type, is_primary, is_foreign, foreign_key_ref, stats):
    """
    根据规则判定字段角色
    
    Args:
        field_name (str): 字段名
        data_type (str): 数据类型
        is_primary (bool): 是否主键
        is_foreign (bool): 是否外键
        foreign_key_ref (str): 外键引用
        stats (dict): 字段统计信息
        
    Returns:
        str: 字段角色
    """
    field_name_lower = field_name.lower()
    data_type_lower = data_type.lower()
    
    # 1. tech_meta - 技术字段，最高优先级
    tech_patterns = [
        r'^etl_', r'^sync_', r'^op_type$', r'^batch_id$', 
        r'^insert_time$', r'^load_', r'^is_deleted$',
        r'^create_time$', r'^update_time$', r'^created_at$', r'^updated_at$',
        r'^create_user$', r'^update_user$', r'^create_by$', r'^update_by$',
        r'^deleted_at$', r'^deleted_by$', r'^deleted$',
        r'^version$', r'^gmt_create$', r'^gmt_modified$',
        r'^sys_', r'^_etl_', r'^_sync_',
        r'^_create_', r'^_update_', r'^_delete_'
    ]
    
    for pattern in tech_patterns:
        if re.search(pattern, field_name_lower):
            return 'tech_meta'
    
    # 2. primary_key - 主键
    unique_rate = (stats['distinct_cnt'] / stats['total_cnt'] * 100) if stats['total_cnt'] > 0 else 0
    null_rate = stats['null_rate']
    
    if is_primary or (unique_rate >= 99.9 and null_rate <= 1.0 and stats['total_cnt'] > 0):
        return 'primary_key'
    
    # 3. foreign_key - 外键
    if is_foreign or foreign_key_ref:
        return 'foreign_key'
    
    # 4. business_time - 业务时间
    time_types = ['timestamp', 'date', 'datetime', 'time']
    is_time_type = any(t in data_type_lower for t in time_types)
    
    if is_time_type:
        time_patterns = [r'_time$', r'_at$', r'_date$']
        exclude_patterns = [r'^create', r'^update', r'^insert', r'^load', r'^gmt_']
        
        # 检查是否是业务时间（排除纯技术时间）
        is_business_time = any(re.search(p, field_name_lower) for p in time_patterns)
        is_tech_time = any(re.search(p, field_name_lower) for p in exclude_patterns)
        
        if is_business_time and not is_tech_time:
            return 'business_time'
    
    # 5. numeric_measure - 数值度量
    numeric_types = ['decimal', 'bigint', 'int', 'float', 'double', 'numeric', 'integer', 'long']
    is_numeric_type = any(t in data_type_lower for t in numeric_types)
    
    if is_numeric_type:
        # 确保不是键、不是技术字段、不是时间
        return 'numeric_measure'
    
    # 6. low_cardinality - 低基数字段
    if (stats['distinct_cnt'] <= 50 and 
        (stats['distinct_rate'] <= 0.5 or stats['total_cnt'] < 10000)):
        return 'low_cardinality'
    
    # 7. business_attr - 兜底角色
    return 'business_attr'


def export_to_single_sheet(metadata_list, output_path, db_type, db_name):
    """
    导出所有表元数据到单个Excel工作表
    
    Args:
        metadata_list (list): 元数据列表
        output_path (str): 输出文件路径
        db_type (str): 数据库类型
        db_name (str): 数据库名称
    """
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "所有表元数据"

    # 定义表头
    headers = ["数据源类型", "业务数据库名称", "表名", "表中文名", "表行数", "字段名", "字段注释", "字段注释填充", "数据类型", "是否为空", "默认值", "主键", "外键", "外键引用", "字段空值率", "字段角色"]

    # 添加表头
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 应用表头样式
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 填充数据
    for item in metadata_list:
        # 处理主键和外键
        is_primary = "YES" if item['key'] == "PRI" else "NO"
        is_foreign = "YES" if item['key'] == "MUL" else "NO"

        # 添加一行数据
        ws.append([
            db_type,
            db_name,
            item['table_name'],
            item.get('table_chinese_name', ""),
            item.get('row_count', 0),
            item['field'],
            item['comment'],
            item['comment_fill'],
            item['type'],
            item['null'],
            item['default'] if item['default'] is not None else "",
            is_primary,
            is_foreign,
            item.get('foreign_key_ref', ""),
            item['fill_rate'],
            item.get('field_role', "")
        ])

    # 调整列宽
    column_widths = [12, 20, 35, 25, 10, 25, 30, 30, 20, 10, 15, 8, 8, 30, 12, 15]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存Excel文件
    wb.save(output_path)
    print(f"✅ 元数据已导出到: {output_path}")


def main():
    """
    主函数
    """
    # 配置名称
    config_name = "yummy_mysql"
    # 构建输出文件路径（项目根目录）
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../..'))
    output_dir = os.path.join(project_root, 'output/metadata_parse')
    output_file = os.path.join(output_dir, "all_tables_metadata.xlsx")

    try:
        # 连接数据库
        print("正在连接数据库...")
        connection, db_config = get_db_connection(config_name)
        db_type = db_config.get('type', 'mysql')
        db_name = db_config.get('dbname', config_name)
        print(f"✅ 数据库连接成功 ({db_type} - {db_name})")

        # 获取所有表
        print("正在获取表列表...")
        tables = get_all_tables(connection)
        print(f"✅ 找到 {len(tables)} 个表")

        # 获取每个表的元数据
        print("正在获取表元数据...")
        metadata_list = []
        for table in tables:
            print(f"  处理表: {table}")
            # 获取表的中文名
            table_chinese_name = get_table_chinese_name(connection, table)
            # 获取表行数
            row_count = get_table_row_count(connection, table)
            # 获取表结构
            columns = get_table_metadata(connection, table)
            # 处理每个字段
            for column in columns:
                field_name = column['Field']
                # 获取字段注释
                comment = get_column_comment(connection, table, field_name)
                # 计算字段注释填充
                if comment:
                    comment_fill = comment
                else:
                    comment_fill = translate_field_name(field_name)
                # 获取字段空值率
                fill_rate = get_column_fill_rate(connection, table, field_name)
                # 获取外键引用信息
                foreign_key_ref = get_foreign_key_reference(connection, table, field_name)
                # 获取字段统计信息
                stats = get_column_stats(connection, table, field_name)
                # 判定字段角色
                is_primary = (column['Key'] == "PRI")
                is_foreign = (column['Key'] == "MUL")
                field_role = determine_field_role(
                    field_name, column['Type'], is_primary, is_foreign, foreign_key_ref, stats
                )

                # 添加到元数据列表
                metadata_list.append({
                    'table_name': table,
                    'table_chinese_name': table_chinese_name,
                    'row_count': row_count,
                    'field': field_name,
                    'comment': comment,
                    'comment_fill': comment_fill,
                    'type': column['Type'],
                    'null': column['Null'],
                    'default': column['Default'],
                    'key': column['Key'],
                    'foreign_key_ref': foreign_key_ref,
                    'fill_rate': fill_rate,
                    'field_role': field_role
                })

        print(f"✅ 共获取 {len(metadata_list)} 个字段的元数据")

        # 导出到Excel
        print("正在导出到Excel...")
        export_to_single_sheet(metadata_list, output_file, db_type, db_name)

        print("\n🎉 操作完成!")

    except Exception as e:
        # 处理异常
        print(f"❌ 操作失败: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭数据库连接
        if 'connection' in locals():
            connection.close()


if __name__ == "__main__":
    # 执行主函数
    main()