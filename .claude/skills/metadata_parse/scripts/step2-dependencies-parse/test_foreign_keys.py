
import os
from openpyxl import load_workbook

# 检查外键信息
all_tables_file = os.path.join('Output', 'all_tables_metadata.xlsx')
if os.path.exists(all_tables_file):
    wb = load_workbook(all_tables_file)
    ws = wb.active
    
    print("检查外键字段：")
    print("=" * 120)
    
    for row in range(2, ws.max_row + 1):
        is_foreign = ws.cell(row=row, column=12).value
        foreign_key_ref = ws.cell(row=row, column=13).value
        
        if is_foreign or foreign_key_ref:
            table_name = ws.cell(row=row, column=3).value
            field_name = ws.cell(row=row, column=5).value
            print(f"表: {table_name}, 字段: {field_name}, 外键: {is_foreign}, 外键引用: {foreign_key_ref}")
    
    wb.close()
