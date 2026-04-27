
import os
from openpyxl import load_workbook

# 检查 t_booking_order 表的字段信息
all_tables_file = os.path.join('Output', 'all_tables_metadata.xlsx')
if os.path.exists(all_tables_file):
    wb = load_workbook(all_tables_file)
    ws = wb.active
    
    print("t_booking_order 表字段信息：")
    print("=" * 120)
    print(f"{'字段名':<30} {'数据类型':<20} {'是否为空':<10} {'默认值':<15} {'主键':<8} {'外键':<8} {'字段注释':<40}")
    print("=" * 120)
    
    for row in range(2, ws.max_row + 1):
        table_name = ws.cell(row=row, column=3).value
        if table_name == 't_booking_order':
            field_name = ws.cell(row=row, column=5).value
            data_type = ws.cell(row=row, column=8).value
            is_null = ws.cell(row=row, column=9).value
            default = ws.cell(row=row, column=10).value
            is_primary = ws.cell(row=row, column=11).value
            is_foreign = ws.cell(row=row, column=12).value
            comment = ws.cell(row=row, column=6).value
            
            data_type_str = data_type if data_type is not None else ""
            comment_str = comment if comment is not None else ""
            print(f"{field_name:<30} {data_type_str:<20} {is_null:<10} {str(default):<15} {is_primary:<8} {is_foreign:<8} {comment_str:<40}")
    
    wb.close()
