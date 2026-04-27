
import os
from openpyxl import load_workbook

# 检查all_tables_metadata.xlsx
print("=" * 80)
print("检查 all_tables_metadata.xlsx")
print("=" * 80)
all_tables_file = os.path.join('Output', 'all_tables_metadata.xlsx')
if os.path.exists(all_tables_file):
    wb = load_workbook(all_tables_file)
    ws = wb.active
    print(f"工作表名称: {ws.title}")
    print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")
    print("\n表头:")
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        headers.append(header)
        print(f"  {col}: {header}")
    print("\n前5行数据:")
    for row in range(2, min(7, ws.max_row + 1)):
        row_data = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            row_data.append(str(value) if value is not None else "")
        print(f"  行 {row}: {row_data}")
    wb.close()

# 检查metadata_field_lineage.xlsx
print("\n" + "=" * 80)
print("检查 metadata_field_lineage.xlsx")
print("=" * 80)
lineage_file = os.path.join('Output', 'metadata_field_lineage.xlsx')
if os.path.exists(lineage_file):
    wb = load_workbook(lineage_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n工作表: {sheet_name}")
        print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")
        if ws.max_row > 0 and ws.max_column > 0:
            print("表头:")
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                headers.append(header)
                print(f"  {col}: {header}")
            print("\n前5行数据:")
            for row in range(2, min(7, ws.max_row + 1)):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    value = ws.cell(row=row, column=col).value
                    row_data.append(str(value) if value is not None else "")
                print(f"  行 {row}: {row_data}")
    wb.close()
