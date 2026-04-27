import os
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import yaml


def load_word_map():
    """
    从Excel参考文档加载词汇映射
    
    Returns:
        dict: 词汇映射字典
    """
    from openpyxl import load_workbook
    
    # 构建词汇映射文件路径（项目根目录）
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../..'))
    word_map_file = os.path.join(project_root, 'Input', 'word_map_reference.xlsx')
    
    word_map = {
        'id': 'ID',
        'name': '名称',
        'code': '代码'
    }
    
    try:
        # 加载Excel文件
        wb = load_workbook(word_map_file)
        ws = wb.active
        
        # 从第二行开始读取数据（第一行是表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[0] and row[1]:  # 确保row不是None且有足够的列
                try:
                    word_map[row[0].lower()] = row[1]
                except Exception:
                    pass
        
    except Exception as e:
        # 如果文件不存在或读取失败，使用默认词汇映射
        print(f"警告: 无法加载词汇映射文件，使用默认映射: {e}")
    
    return word_map


def split_field_name(field_name):
    """
    拆解字段名
    
    Args:
        field_name (str): 字段名
        
    Returns:
        list: 拆解后的字段名部分列表
    """
    # 处理下划线分隔的字段名
    if '_' in field_name:
        parts = field_name.split('_')
        return [part.lower() for part in parts if part]
    
    # 处理驼峰命名和连续小写字母的字段名
    parts = []
    current_part = ''
    
    # 常见词汇列表，用于智能拆解字段名
    common_words = [
        'updated', 'created', 'deleted', 'enabled', 'disabled', 'visible', 'hidden',
        'shared', 'personal', 'global', 'local', 'default', 'custom', 'active', 'inactive',
        'classification', 'telephone', 'location', 'department', 'organization',
        'information', 'description', 'institution', 'administration', 'implementation',
        'configuration', 'documentation', 'communication', 'relationship', 'development',
        'management', 'performance', 'environment', 'infrastructure', 'authentication',
        'company', 'office', 'parent', 'depart', 'address', 'version', 'history', 'archive',
        'public', 'private',
        'update', 'create', 'delete', 'status', 'term', 'date', 'job', 'user',
        'name', 'code', 'type', 'time', 'id', 'number', 'email', 'phone',
        'url', 'path', 'size', 'cost', 'price', 'profit', 'value', 'count',
        'total', 'amount', 'unit', 'level', 'rank', 'order', 'product', 'service',
        'project', 'task', 'event', 'file', 'image', 'video', 'audio', 'role',
        'access', 'security', 'backup', 'restore', 'log', 'audit', 'alert', 'message',
        'comment', 'review', 'approval', 'reject', 'accept', 'submit', 'cancel', 'confirm',
        'verify', 'login', 'logout', 'register', 'reset', 'change', 'edit', 'view',
        'list', 'detail', 'report', 'dashboard', 'statistics', 'analysis', 'forecast',
        'trend', 'pattern', 'model', 'template', 'format', 'standard', 'rule', 'policy'
    ]
    
    # 去重并按长度降序排序，确保长词汇优先匹配
    unique_words = sorted(list(set(common_words)), key=len, reverse=True)
    
    i = 0
    n = len(field_name)
    
    while i < n:
        char = field_name[i]
        
        # 处理大写字母（驼峰命名）
        if char.isupper():
            if current_part:
                parts.append(current_part)
            current_part = char.lower()
            i += 1
        # 处理数字
        elif char.isdigit():
            if current_part and not current_part[-1].isdigit():
                parts.append(current_part)
                current_part = char
            else:
                current_part += char
            i += 1
        # 处理特殊字符
        elif not char.isalnum():
            if current_part:
                parts.append(current_part)
                current_part = ''
            i += 1
        # 处理小写字母
        else:
            matched = False
            # 尝试匹配最长的词汇
            for j in range(min(i + 20, n), i, -1):
                substring = field_name[i:j].lower()
                if substring in unique_words:
                    if current_part:
                        parts.append(current_part)
                        current_part = ''
                    parts.append(substring)
                    i = j
                    matched = True
                    break
            
            # 如果没有匹配到词汇，继续累积
            if not matched:
                current_part += char.lower()
                i += 1
    
    # 添加最后一个部分
    if current_part:
        parts.append(current_part)
    
    # 如果没有拆分成任何部分，返回原字段名的小写形式
    if not parts:
        parts = [field_name.lower()]
    
    return parts


def translate_field_name(field_name):
    """
    翻译字段名为中文
    
    Args:
        field_name (str): 字段名
        
    Returns:
        str: 翻译后的中文名称
    """
    # 加载词汇映射
    word_map = load_word_map()
    
    # 拆解字段名
    parts = split_field_name(field_name)
    
    # 翻译每个部分
    translated_parts = []
    for part in parts:
        if part in word_map:
            translated_parts.append(word_map[part])
        else:
            # 对于未识别的部分，保留原词
            translated_parts.append(part)
    
    # 组合翻译结果
    return ''.join(translated_parts)


def parse_table_structure(html_content):
    """
    解析表结构
    
    Args:
        html_content (str): HTML内容
        
    Returns:
        dict: 表结构信息
    """
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 提取表名
        table_name = ""
        # 查找所有h2标签，找到格式为"表t_table_name"的标题
        h2_tags = soup.find_all('h2')
        for h2 in h2_tags:
            if h2 and h2.text:
                # 匹配"表t_table_name"格式，确保只匹配以t_开头的表名
                match = re.search(r'表(t_\w+)', h2.text)
                if match:
                    table_name = match.group(1)
                    break
        
        # 如果没有找到表名，尝试从h3标签中提取（处理Stan11_1.htm这种情况）
        if not table_name:
            h3_tags = soup.find_all('h3')
            for h3 in h3_tags:
                if h3 and h3.text:
                    # 匹配"表t_table_name的栏的清单"格式
                    match = re.search(r'表(t_\w+)的栏的清单', h3.text)
                    if match:
                        table_name = match.group(1)
                        break
        
        if not table_name:
            return {'table_name': '', 'table_chinese_name': '', 'fields': []}
        
        # 提取表注释（从代码预览部分）
        table_comment = ""
        # 尝试从代码预览中提取表DDL和注释
        # 查找包含"代码预览"的h3标签，特别是"表的代码预览table_name"格式
        code_preview = None
        h3_tags = soup.find_all('h3')
        for h3 in h3_tags:
            if h3 and h3.text and ('代码预览' in h3.text):
                code_preview = h3
                break
        
        code_text = ""
        if code_preview:
            code_table = code_preview.find_next('table')
            if code_table:
                code_text = code_table.text
                if code_text:
                    # 从ALTER TABLE语句中提取表注释
                    match = re.search(r"alter table .*? comment '(.*?)';", code_text, re.DOTALL | re.IGNORECASE)
                    if match:
                        table_comment = match.group(1)
        
        # 如果从代码预览中没有找到表注释，尝试从表的卡片信息中提取
        if not table_comment:
            table_card = soup.find('h3', string=lambda text: text and '卡片' in text)
            if table_card:
                form_table = table_card.find_next('table', class_='Form')
                if form_table:
                    for tr in form_table.find_all('tr'):
                        tds = tr.find_all('td')
                        if len(tds) >= 2:
                            key = tds[0].text.strip()
                            value = tds[1].text.strip()
                            if '名称' in key and value and value != table_name:
                                # 如果名称字段的值不是表名，可能是中文名称
                                table_comment = value
                                break
        
        # 提取字段信息
        fields = []
        # 查找字段列表部分
        # 查找所有h3标签，找到包含"栏的清单"或"列的清单"的部分
        h3_tags = soup.find_all('h3')
        columns_section = None
        for h3 in h3_tags:
            if h3 and h3.text and ('栏的清单' in h3.text or '列的清单' in h3.text):
                columns_section = h3
                break
        
        if columns_section:
            columns_table = columns_section.find_next('table')
            if columns_table:
                rows = columns_table.find_all('tr')
                if rows:
                    rows = rows[1:]  # 跳过表头
                    field_names = []
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) >= 2:
                            field_name = cells[1].text.strip()  # 代码列包含字段名
                            if field_name:
                                field_names.append(field_name)
                    
                    # 为每个字段提取详细信息
                    for field_name in field_names:
                        # 查找字段的详细信息部分
                        field_section = None
                        # 查找所有h2标签，找到包含字段名的部分
                        h2_tags = soup.find_all('h2')
                        for h2 in h2_tags:
                            if h2 and h2.text and field_name in h2.text:
                                field_section = h2
                                break
                        
                        # 提取数据类型、是否强制等信息
                        field_info = {}
                        if field_section:
                            field_card = field_section.find_next('h3')
                            if field_card:
                                form_table = field_card.find_next('table', class_='Form')
                                if form_table:
                                    for tr in form_table.find_all('tr'):
                                        tds = tr.find_all('td')
                                        if len(tds) >= 2:
                                            key = tds[0].text.strip()
                                            value = tds[1].text.strip()
                                            if '数据类型' in key:
                                                field_info['type'] = value
                                            elif '强制' in key:
                                                field_info['null'] = "NO" if value == "TRUE" else "YES"
                        
                        # 提取默认值
                        if field_section:
                            validation_section = field_section.find_next('h3', string=lambda text: text and '验证' in text)
                            if validation_section:
                                validation_table = validation_section.find_next('table', class_='Form')
                                if validation_table:
                                    for tr in validation_table.find_all('tr'):
                                        tds = tr.find_all('td')
                                        if len(tds) >= 2:
                                            key = tds[0].text.strip()
                                            value = tds[1].text.strip()
                                            if '默认值' in key and value and value != '&nbsp;':
                                                field_info['default'] = value
                        
                        # 从代码预览中的CREATE TABLE语句提取字段信息
                        comment = ""
                        if code_text:
                            # 从CREATE TABLE语句中提取字段信息
                            try:
                                # 直接在整个代码文本中查找字段定义
                                # 处理包含括号的类型（如 decimal(18,2)）
                                # 更精确的正则表达式，避免匹配到注释部分
                                pattern = rf"{field_name}\s+(?:national\s+)?([a-zA-Z]+(?:\([^)]*\))?)(?:\s+(?:not\s+null|default\s+[^,]+))?(?:\s+comment\s+['\"]([\s\S]*?)['\"])?(?:,|$)"
                                match = re.search(pattern, code_text, re.IGNORECASE)
                                if match:
                                    # 提取数据类型
                                    data_type = match.group(1).strip()
                                    # 过滤掉无效的类型
                                    if data_type not in ['on', 'drop', 'create', 'alter', 'primary', 'key', 'index']:
                                        field_info['type'] = data_type
                                    # 提取注释
                                    if match.group(2):
                                        comment = match.group(2).strip()
                                
                                # 如果没有找到注释，尝试更简单的模式
                                if not comment:
                                    simple_pattern = rf"{field_name}.*?comment\s+['\"]([\s\S]*?)['\"]"
                                    simple_match = re.search(simple_pattern, code_text, re.IGNORECASE)
                                    if simple_match:
                                        comment = simple_match.group(1).strip()
                            except Exception as e:
                                pass
                        
                        # 从字段的详细信息部分提取注释
                        if not comment and field_section:
                            try:
                                field_card = field_section.find_next('h3')
                                if field_card:
                                    form_table = field_card.find_next('table', class_='Form')
                                    if form_table:
                                        for tr in form_table.find_all('tr'):
                                            tds = tr.find_all('td')
                                            if len(tds) >= 2:
                                                key = tds[0].text.strip()
                                                value = tds[1].text.strip()
                                                if '注释' in key and value:
                                                    comment = value
                                                    break
                            except Exception as e:
                                pass
                        
                        # 计算字段注释填充
                        if comment:
                            comment_fill = comment
                        else:
                            try:
                                comment_fill = translate_field_name(field_name)
                            except Exception:
                                comment_fill = field_name
                        
                        # 构建字段信息
                        field_data = {
                            'field': field_name,
                            'type': field_info.get('type', ''),
                            'null': field_info.get('null', 'YES'),
                            'default': field_info.get('default', ''),
                            'comment': comment,
                            'comment_fill': comment_fill
                        }
                        fields.append(field_data)
        
        # 提取主键信息
        primary_keys = []
        # 从代码预览中的PRIMARY KEY语句提取
        if code_text:
            # 提取PRIMARY KEY后面的字段名
            match = re.search(r'primary key \(([^\)]+)\)', code_text, re.IGNORECASE)
            if match:
                pk_fields = match.group(1).split(',')
                primary_keys = [field.strip() for field in pk_fields]
        
        # 提取外键信息
        foreign_keys = {}
        # 从代码预览中的外键定义提取
        if code_text:
            # 提取外键定义
            try:
                fk_patterns = re.findall(r'alter table .*? add constraint .*? foreign key \(([^\)]+)\) references ([^\(]+)\(([^\)]+)\)', code_text, re.DOTALL | re.IGNORECASE)
                for fk in fk_patterns:
                    if len(fk) >= 3:
                        fk_field = fk[0].strip()
                        ref_table = fk[1].strip()
                        ref_field = fk[2].strip()
                        # 处理表名中的模式名（如zhaoxiang.t_table）
                        ref_table = ref_table.split('.')[-1]  # 取最后一部分作为表名
                        foreign_keys[fk_field] = f"{ref_table}.{ref_field}"
            except Exception:
                pass
        
        # 为字段添加主键和外键信息
        for field in fields:
            field_name = field['field']
            field['key'] = "PRI" if field_name in primary_keys else ("MUL" if field_name in foreign_keys else "")
            field['foreign_key_ref'] = foreign_keys.get(field_name, "")
            # 模拟空值率（PowerDesigner报告中没有实际数据）
            field['fill_rate'] = "0.00%"  # 默认为0%
        
        return {
            'table_name': table_name,
            'table_chinese_name': table_comment,
            'fields': fields
        }
    except Exception as e:
        print(f"解析表结构时出错: {e}")
        return {'table_name': '', 'table_chinese_name': '', 'fields': []}


def parse_powerdesigner_report(report_dir):
    """
    解析PowerDesigner报告
    
    Args:
        report_dir (str): 报告目录路径
        
    Returns:
        list: 元数据列表
    """
    metadata_list = []
    
    # 找到所有Stan*.htm文件
    files_dir = os.path.join(report_dir, 'Standard Physical Report_files')
    if not os.path.exists(files_dir):
        print(f"错误: 报告文件目录不存在: {files_dir}")
        return metadata_list
    
    # 获取所有Stan*.htm文件
    import glob
    htm_files = glob.glob(os.path.join(files_dir, 'Stan*.htm'))
    
    # 按文件名排序，确保 Stan11.htm 在 Stan11_1.htm 之前处理
    htm_files.sort()
    
    # 解析每个htm文件
    for file_path in htm_files:
        # 跳过处理拆分的文件（如 Stan11_1.htm）
        if '_' in os.path.basename(file_path) and file_path.endswith('.htm'):
            basename = os.path.basename(file_path)
            if re.search(r'\d+_\d+\.htm$', basename):
                print(f"跳过处理拆分文件: {basename}")
                continue
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            table_info = parse_table_structure(html_content)
            if table_info['table_name'] and table_info['fields']:
                print(f"解析表: {table_info['table_name']}")
                for field in table_info['fields']:
                    metadata_list.append({
                        'table_name': table_info['table_name'],
                        'table_chinese_name': table_info['table_chinese_name'],
                        'field': field['field'],
                        'comment': field['comment'],
                        'comment_fill': field['comment_fill'],
                        'type': field['type'],
                        'null': field['null'],
                        'default': field['default'],
                        'key': field['key'],
                        'foreign_key_ref': field['foreign_key_ref'],
                        'fill_rate': field['fill_rate']
                    })
        except Exception as e:
            print(f"解析文件 {file_path} 时出错: {e}")
    
    return metadata_list

def export_to_single_sheet(metadata_list, output_path, db_type, db_name):
    """
    导出所有表元数据到单个Excel工作表
    
    Args:
        metadata_list (list): 元数据列表
        output_path (str): 输出文件路径
        db_type (str): 数据库类型
        db_name (str): 数据库名称
    """
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "所有表元数据"
    
    # 定义表头
    headers = ["数据源类型", "业务数据库名称", "表名", "表中文名", "字段名", "字段注释", "字段注释填充", "数据类型", "是否为空", "默认值", "主键", "外键", "外键引用", "字段空值率"]
    
    # 添加表头
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 应用表头样式
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 填充数据
    for item in metadata_list:
        # 处理主键和外键
        is_primary = "YES" if item['key'] == "PRI" else "NO"
        is_foreign = "YES" if item['key'] == "MUL" else "NO"
        
        # 添加一行数据
        ws.append([
            db_type,
            db_name,
            item['table_name'],
            item.get('table_chinese_name', ""),
            item['field'],
            item['comment'],
            item['comment_fill'],
            item['type'],
            item['null'],
            item['default'] if item['default'] is not None else "",
            is_primary,
            is_foreign,
            item.get('foreign_key_ref', ""),
            item['fill_rate']
        ])
    
    # 调整列宽
    column_widths = [12, 20, 35, 25, 25, 30, 30, 20, 10, 15, 8, 8, 30, 12]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存Excel文件
    wb.save(output_path)
    print(f"✅ 元数据已导出到: {output_path}")


def main():
    """
    主函数
    """
    # 构建路径
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../..'))
    report_dir = os.path.join(project_root, 'Input', 'application', 'Report')
    output_dir = os.path.join(project_root, 'Output')
    output_file = os.path.join(output_dir, "all_tables_metadata.xlsx")
    
    try:
        # 解析PowerDesigner报告
        print("正在解析PowerDesigner报告...")
        metadata_list = parse_powerdesigner_report(report_dir)
        print(f"✅ 共解析 {len(metadata_list)} 个字段的元数据")
        
        # 导出到Excel
        print("正在导出到Excel...")
        export_to_single_sheet(metadata_list, output_file, "powerdesigner", "PhysicalDataModel_1")
        
        print("\n🎉 操作完成!")
        
    except Exception as e:
        # 处理异常
        print(f"❌ 操作失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 执行主函数
    main()
