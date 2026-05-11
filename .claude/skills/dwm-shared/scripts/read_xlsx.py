#!/usr/bin/env python3
"""
Excel (xlsx) reader for dwm sub-skills (shared tool).

Reads upstream metadata xlsx files and returns list of dicts,
providing a consistent interface similar to read_csv.

Usage:
    # Read all rows from default sheet
    python read_xlsx.py output/metadata_parse/all_tables_metadata.xlsx

    # Filter by column value
    python read_xlsx.py output/metadata_parse/all_tables_metadata.xlsx --where "字段角色=foreign_key"

    # Select specific columns
    python read_xlsx.py output/metadata_parse/all_tables_metadata.xlsx --select "表名,字段名,字段角色"

    # Count rows
    python read_xlsx.py output/metadata_parse/all_tables_metadata.xlsx --where "字段角色=foreign_key" --count

    # Distinct values of a column
    python read_xlsx.py output/metadata_parse/all_tables_metadata.xlsx --distinct "字段角色"

    # Specify sheet name (default: first sheet)
    python read_xlsx.py output/ods_generator/all_tables_metadata_ods.xlsx --sheet "Sheet1"
"""

import json
import sys
import argparse

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def read_xlsx(filepath: str, sheet_name: str | None = None) -> list[dict]:
    """Read xlsx file, return list of dicts (first row as headers)."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = "" if val is None else str(val)
        result.append(record)

    return result


def apply_filters(rows: list[dict], filters: list[str]) -> list[dict]:
    """Filter rows by key=value conditions (AND logic)."""
    for f in filters:
        key, value = f.split("=", 1)
        rows = [r for r in rows if r.get(key, "") == value]
    return rows


def select_columns(rows: list[dict], columns: list[str]) -> list[dict]:
    """Keep only specified columns."""
    return [{k: r.get(k, "") for k in columns} for r in rows]


def main():
    parser = argparse.ArgumentParser(description="Read xlsx, output JSON")
    parser.add_argument("filepath", help="Input xlsx file path")
    parser.add_argument("--sheet", help="Sheet name (default: active sheet)")
    parser.add_argument("--where", "-w", action="append", default=[], help="Filter: key=value (repeatable, AND)")
    parser.add_argument("--select", "-s", help="Comma-separated columns to include")
    parser.add_argument("--count", action="store_true", help="Output row count only")
    parser.add_argument("--distinct", "-d", help="Output distinct values of a column")
    args = parser.parse_args()

    rows = read_xlsx(args.filepath, args.sheet)

    if args.where:
        rows = apply_filters(rows, args.where)

    if args.distinct:
        values = sorted(set(r.get(args.distinct, "") for r in rows))
        print(json.dumps(values, ensure_ascii=False, indent=2))
        return

    if args.count:
        print(len(rows))
        return

    if args.select:
        rows = select_columns(rows, args.select.split(","))

    print(json.dumps(rows, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
