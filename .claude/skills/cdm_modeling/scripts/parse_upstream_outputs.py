"""
parse_upstream_outputs.py - DWM 上游建设规格解析模块
====================================================
读取 dwm-business-process、dwm-dimension、dwm-matrix 的标准产物，
形成 CDM DIM/DWD 生成所需的统一建模上下文。
"""

import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List


class UpstreamOutputParser:
    """解析最新 DWM CSV/XLSX 交付物，并兼容 cdm-modeling 旧列名。"""

    MATRIX_FIXED_COLUMNS = {
        "主题域",
        "业务过程",
        "业务过程代码",
        "粒度声明",
        "事实表名称",
        "事实表类型",
    }

    def __init__(
        self,
        business_process_file: str = "",
        subject_area_file: str = "",
        dim_spec_file: str = "",
        dwd_fact_spec_file: str = "",
        bus_matrix_file: str = "",
        dim_join_spec_file: str = "",
        dwd_join_spec_file: str = "",
        base_dir: Path = Path("."),
        logger=None,
    ):
        self.base_dir = Path(base_dir)
        self.business_process_file = self._optional_path(business_process_file)
        self.subject_area_file = self._optional_path(subject_area_file)
        self.dim_spec_file = self._optional_path(dim_spec_file)
        self.dwd_fact_spec_file = self._optional_path(dwd_fact_spec_file)
        self.bus_matrix_file = self._optional_path(bus_matrix_file)
        self.dim_join_spec_file = self._optional_path(dim_join_spec_file)
        self.dwd_join_spec_file = self._optional_path(dwd_join_spec_file)
        self.logger = logger
        self.warnings: List[str] = []

    def parse(self) -> Dict[str, Any]:
        business_rows = self._load_configured(self.business_process_file, "业务过程清单")
        subject_rows = self._load_configured(self.subject_area_file, "主题域注册表")
        dim_rows = self._load_configured(self.dim_spec_file, "DIM 建设清单", required=True)
        dwd_rows = self._load_configured(self.dwd_fact_spec_file, "DWD 建设清单", required=True)
        matrix_rows = self._load_configured(self.bus_matrix_file, "总线矩阵", required=True)
        dim_join_rows = self._load_configured(self.dim_join_spec_file, "DIM ODS 关联清单")
        dwd_join_rows = self._load_configured(self.dwd_join_spec_file, "DWD ODS 关联清单")

        self._validate_input_columns(dim_rows, dwd_rows, matrix_rows)

        dimensions = self._extract_dimensions_from_dim_spec(dim_rows, dim_join_rows)
        matrix_links = self._extract_matrix_links(matrix_rows, dimensions)
        processes = self._extract_processes_from_dwd_spec(
            dwd_rows=dwd_rows,
            dimensions=dimensions,
            matrix_links=matrix_links,
            business_rows=business_rows,
            subject_rows=subject_rows,
            dwd_join_rows=dwd_join_rows,
        )

        model = {
            "processes": processes,
            "dimensions": dimensions,
            "business_processes": business_rows,
            "subject_areas": subject_rows,
            "matrix_links": matrix_links,
            "ods_tables": self._extract_lineage_tables(dim_rows, dwd_rows),
            "warnings": self.warnings,
            "input_contract": {
                "business_process_file": self._path_text(self.business_process_file),
                "subject_area_file": self._path_text(self.subject_area_file),
                "dim_spec_file": self._path_text(self.dim_spec_file),
                "dwd_fact_spec_file": self._path_text(self.dwd_fact_spec_file),
                "bus_matrix_file": self._path_text(self.bus_matrix_file),
                "dim_join_spec_file": self._path_text(self.dim_join_spec_file),
                "dwd_join_spec_file": self._path_text(self.dwd_join_spec_file),
            },
        }
        self._validate(model)
        return model

    def _optional_path(self, raw_path: str) -> Path | None:
        return self._resolve_path(raw_path) if raw_path else None

    def _resolve_path(self, raw_path: str) -> Path:
        path = Path(raw_path)
        return path if path.is_absolute() else self.base_dir / path

    def _path_text(self, path: Path | None) -> str:
        return str(path) if path else ""

    def _load_configured(self, path: Path | None, label: str, required: bool = False) -> List[Dict[str, str]]:
        if path is None:
            if required:
                self._warn(f"缺少必需输入配置: {label}")
            return []
        if not path.exists():
            self._warn(f"输入文件不存在: {label} ({path})")
            return []
        return self._load_table(path)

    def _load_table(self, path: Path) -> List[Dict[str, str]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv(path)
        if suffix == ".xlsx":
            return self._load_xlsx(path)
        self._warn(f"不支持的输入格式: {path}，仅支持 .csv 和 .xlsx")
        return []

    def _load_csv(self, path: Path) -> List[Dict[str, str]]:
        with open(path, "r", encoding="utf-8-sig", newline="") as file:
            return [self._clean_row(row) for row in csv.DictReader(file)]

    def _load_xlsx(self, path: Path) -> List[Dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self._warn("读取 XLSX 需要安装 openpyxl")
            return []

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [self._stringify(value) for value in next(rows)]
        except StopIteration:
            workbook.close()
            return []

        result: List[Dict[str, str]] = []
        for values in rows:
            if not any(value is not None and str(value).strip() for value in values):
                continue
            row = {
                headers[index]: self._stringify(value)
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            result.append(self._clean_row(row))
        workbook.close()
        return result

    def _validate_input_columns(
        self,
        dim_rows: List[Dict[str, str]],
        dwd_rows: List[Dict[str, str]],
        matrix_rows: List[Dict[str, str]],
    ) -> None:
        self._require_aliases(
            "dwm_dim_spec",
            dim_rows,
            {
                "维度表名": ["维度表名", "DIM表名", "dim_table_name"],
                "字段名": ["字段名", "DIM字段名", "dim_column_name"],
                "字段角色": ["字段角色", "column_role"],
                "来源ODS表": ["来源ODS表", "ods_table_name"],
                "来源ODS字段": ["来源ODS字段", "ods_column_name"],
                "ODS数据类型": ["ODS数据类型", "ODS字段数据类型", "ods_data_type"],
            },
        )
        self._require_aliases(
            "dwm_dwd_fact_spec",
            dwd_rows,
            {
                "DWD表名": ["DWD表名", "dwd_table_name"],
                "业务过程标准名": ["业务过程标准名", "bp_standard_name"],
                "字段名": ["字段名", "DWD字段名", "dwd_column_name"],
                "字段角色": ["字段角色", "column_role"],
                "来源ODS表": ["来源ODS表", "ods_table_name"],
                "来源ODS字段": ["来源ODS字段", "ods_column_name"],
                "ODS数据类型": ["ODS数据类型", "ODS字段数据类型", "ods_data_type"],
            },
        )
        self._require_aliases(
            "dwm_bus_matrix",
            matrix_rows,
            {
                "业务过程代码": ["业务过程代码", "业务过程标准名", "bp_standard_name"],
                "事实表名称": ["事实表名称", "DWD表名", "dwd_table_name"],
            },
        )

    def _require_aliases(
        self,
        label: str,
        rows: List[Dict[str, str]],
        requirements: Dict[str, List[str]],
    ) -> None:
        if not rows:
            return
        headers = set(rows[0].keys())
        missing = [name for name, aliases in requirements.items() if not headers.intersection(aliases)]
        if missing:
            self._warn(f"{label} 缺少必要列: {', '.join(missing)}")

    def _extract_dimensions_from_dim_spec(
        self,
        rows: List[Dict[str, str]],
        join_rows: List[Dict[str, str]],
    ) -> List[Dict[str, Any]]:
        grouped: Dict[str, List[Dict[str, str]]] = {}
        for row in rows:
            table_name = self._pick(row, ["维度表名", "DIM表名", "dim_table_name"]).strip()
            if table_name:
                grouped.setdefault(table_name, []).append(row)

        joins_by_table = self._group_join_specs(join_rows, ["维度表名", "DIM表名", "dim_table_name"])
        dimensions: List[Dict[str, Any]] = []
        for table_name, table_rows in grouped.items():
            first = table_rows[0]
            entity = self._entity_from_dim_table(table_name)
            display_name = self._pick(first, ["维度中文名称", "维度中文名", "dimension_name"]) or table_name
            business_key = ""
            business_key_source = ""
            business_key_type = "STRING"
            attributes: List[Dict[str, Any]] = []
            field_scd_types: List[int] = []

            for row in sorted(table_rows, key=self._sort_order):
                field_name = self._pick(row, ["字段名", "DIM字段名", "dim_column_name"]).strip()
                if not field_name:
                    continue
                role = self._pick(row, ["字段角色", "column_role"]).strip().lower()
                field_scd = self._pick(row, ["SCD类型", "scd_type"])
                if field_scd and field_scd != "-":
                    field_scd_types.append(self._parse_scd_type(field_scd))
                if role in {"bk", "business_key", "pk"}:
                    if not business_key:
                        business_key = field_name
                        business_key_source = self._pick(row, ["来源ODS字段", "ods_column_name"]) or field_name
                        business_key_type = self._normalize_data_type(
                            self._pick(row, ["ODS数据类型", "ODS字段数据类型", "ods_data_type"])
                        )
                    continue
                if role == "attribute":
                    attributes.append({
                        "name": field_name,
                        "source_field": self._pick(row, ["来源ODS字段", "ods_column_name"]) or field_name,
                        "type": self._normalize_data_type(
                            self._pick(row, ["ODS数据类型", "ODS字段数据类型", "ods_data_type"])
                        ),
                        "description": self._pick(row, ["字段中文说明", "dim_column_comment"]) or field_name,
                        "scd_type": self._parse_scd_type(field_scd),
                    })

            if not business_key:
                business_key = f"{entity}_id"
                business_key_source = business_key
                self._warn(f"DIM {table_name} 未找到 pk/bk 字段，使用默认业务键: {business_key}")

            table_scd = self._pick(first, ["SCD策略", "table_scd_type", "default_scd_type"])
            scd_type = self._parse_scd_type(table_scd) if table_scd else max(field_scd_types or [1])
            source_tables = self._unique_values(table_rows, ["来源ODS表", "ods_table_name"])
            source_joins = joins_by_table.get(table_name, [])
            source_tables = self._merge_unique(source_tables, [item["source_table"] for item in source_joins])

            dimensions.append({
                "table_name": table_name,
                "name": display_name,
                "description": self._pick(first, ["维度描述", "dimension_description"]),
                "entity": entity,
                "business_key": business_key,
                "business_key_source": business_key_source,
                "business_key_type": business_key_type,
                "attributes": self._dedupe_fields(attributes),
                "source_tables": source_tables,
                "source_joins": source_joins,
                "scd_type": scd_type,
                "is_conformed": self._pick(first, ["是否一致性维度", "is_conformed_dimension"]),
                "shared_scope": self._pick(first, ["跨事实表共享范围", "shared_business_processes"]),
            })
        return dimensions

    def _extract_matrix_links(
        self,
        rows: List[Dict[str, str]],
        dimensions: List[Dict[str, Any]],
    ) -> Dict[str, List[str]]:
        header_to_dim: Dict[str, str] = {}
        for dim in dimensions:
            table_name = dim["table_name"]
            names = {
                table_name,
                dim.get("name", ""),
                self._normalize_name(table_name),
                self._normalize_name(dim.get("name", "")),
            }
            for name in names:
                if name:
                    header_to_dim[name] = table_name

        links: Dict[str, List[str]] = {}
        for row in rows:
            table_name = self._pick(row, ["事实表名称", "DWD表名", "dwd_table_name"])
            process_name = self._pick(row, ["业务过程代码", "业务过程标准名", "bp_standard_name"])
            linked_dims: List[str] = []
            for header, marker in row.items():
                if header in self.MATRIX_FIXED_COLUMNS or not self._is_positive_marker(marker):
                    continue
                dim_table = header_to_dim.get(header) or header_to_dim.get(self._normalize_name(header))
                if dim_table and dim_table not in linked_dims:
                    linked_dims.append(dim_table)
                elif not dim_table:
                    self._warn(f"总线矩阵维度列无法映射到 dwm_dim_spec: {header}")
            if table_name:
                links[table_name] = linked_dims
            if process_name:
                links[f"bp:{process_name}"] = linked_dims
        return links

    def _extract_processes_from_dwd_spec(
        self,
        dwd_rows: List[Dict[str, str]],
        dimensions: List[Dict[str, Any]],
        matrix_links: Dict[str, List[str]],
        business_rows: List[Dict[str, str]],
        subject_rows: List[Dict[str, str]],
        dwd_join_rows: List[Dict[str, str]],
    ) -> List[Dict[str, Any]]:
        grouped: Dict[str, List[Dict[str, str]]] = {}
        for row in dwd_rows:
            table_name = self._pick(row, ["DWD表名", "dwd_table_name"]).strip()
            if table_name:
                grouped.setdefault(table_name, []).append(row)

        business_by_name = {
            self._pick(row, ["业务过程英文名称", "业务过程标准名", "bp_standard_name"]): row
            for row in business_rows
            if self._pick(row, ["业务过程英文名称", "业务过程标准名", "bp_standard_name"])
        }
        valid_subject_codes = {
            self._pick(row, ["主题域编码", "subject_area_code"])
            for row in subject_rows
            if self._pick(row, ["主题域编码", "subject_area_code"])
        }
        joins_by_table = self._group_join_specs(dwd_join_rows, ["DWD表名", "dwd_table_name"])

        processes: List[Dict[str, Any]] = []
        for table_name, table_rows in grouped.items():
            first = table_rows[0]
            process_name = self._pick(first, ["业务过程标准名", "bp_standard_name"])
            business_row = business_by_name.get(process_name, {})
            domain = self._pick(first, ["主题域编码", "subject_area_code"]) or self._pick(
                business_row, ["主题域编码", "subject_area_code"]
            )
            grain = self._pick(first, ["粒度声明", "grain_statement"]) or self._pick(
                business_row, ["粒度声明", "grain_statement"]
            )
            fact_type = self._pick(first, ["事实表类型", "fact_type"]) or self._pick(
                business_row, ["事实表类型", "fact_type"]
            ) or "transaction"

            fields: List[Dict[str, Any]] = []
            measures: List[Dict[str, Any]] = []
            explicit_refs: List[Dict[str, str]] = []
            grain_keys: List[str] = []

            for row in sorted(table_rows, key=self._sort_order):
                role = self._pick(row, ["字段角色", "column_role"]).strip().lower()
                field_name = self._pick(row, ["字段名", "DWD字段名", "dwd_column_name"]).strip()
                if not field_name:
                    continue
                source_field = self._pick(row, ["来源ODS字段", "ods_column_name"]) or field_name
                field = {
                    "name": field_name,
                    "source_field": source_field,
                    "source_table": self._pick(row, ["来源ODS表", "ods_table_name"]),
                    "type": self._normalize_data_type(
                        self._pick(row, ["ODS数据类型", "ODS字段数据类型", "ods_data_type"])
                    ),
                    "description": self._pick(row, ["字段中文说明", "dwd_column_comment"]) or field_name,
                    "role": role,
                    "sort_order": self._sort_order(row),
                }
                fields.append(field)
                if role == "grain_key":
                    grain_keys.append(field_name)
                if role == "measure":
                    measures.append({
                        **field,
                        "aggregation": (self._pick(row, ["聚合建议", "agg_suggest"]) or "SUM").upper(),
                        "measure_type": self._pick(row, ["度量类型", "measure_type"]),
                        "unit": self._pick(row, ["度量单位", "unit"]),
                        "is_derived": self._pick(row, ["是否派生", "is_derived"]),
                        "derived_logic": self._pick(row, ["派生逻辑", "derived_logic"]),
                    })
                dim_table = self._pick(row, ["关联DIM表", "ref_dim_table"]).strip()
                if dim_table:
                    explicit_refs.append(self._make_dimension_ref(row, field, dim_table, dimensions))

            if not grain_keys:
                declared_keys = self._pick(business_row, ["粒度键", "grain_key"])
                grain_keys = self._split_values(declared_keys)

            linked_dims = matrix_links.get(table_name, matrix_links.get(f"bp:{process_name}", []))
            dimension_refs = self._resolve_dimension_refs(
                table_name=table_name,
                fields=fields,
                dimensions=dimensions,
                linked_dim_tables=linked_dims,
                explicit_refs=explicit_refs,
                matrix_available=bool(matrix_links),
            )
            source_joins = joins_by_table.get(table_name, [])
            source_tables = self._unique_values(table_rows, ["来源ODS表", "ods_table_name"])
            source_tables = self._merge_unique(source_tables, [item["source_table"] for item in source_joins])
            if not source_tables:
                source_tables = self._split_values(self._pick(business_row, ["涉及ODS表", "ods_tables"]))

            if valid_subject_codes and domain and domain not in valid_subject_codes:
                self._warn(f"DWD {table_name} 引用了未注册主题域: {domain}")

            processes.append({
                "table_name": table_name,
                "domain": domain,
                "business_process": process_name,
                "display_name": self._pick(first, ["事实表中文名称", "fact_table_name_cn"])
                or self._pick(business_row, ["业务过程中文名称", "business_process_name_cn"]),
                "grain": grain,
                "grain_keys": grain_keys,
                "business_key": grain_keys[0] if grain_keys else "",
                "fact_type": fact_type,
                "dimensions": [ref["entity"] for ref in dimension_refs],
                "dimensions_authoritative": bool(linked_dims),
                "dimension_refs": dimension_refs,
                "fields": self._dedupe_fields(fields),
                "detail_fields": self._dedupe_fields(
                    [field for field in fields if field["role"] in {"degenerate_dim", "low_card_attr", "business_time"}]
                ),
                "measures": self._dedupe_fields(measures),
                "source_tables": source_tables,
                "source_joins": source_joins,
            })
        return processes

    def _resolve_dimension_refs(
        self,
        table_name: str,
        fields: List[Dict[str, Any]],
        dimensions: List[Dict[str, Any]],
        linked_dim_tables: List[str],
        explicit_refs: List[Dict[str, str]],
        matrix_available: bool,
    ) -> List[Dict[str, str]]:
        refs = list(explicit_refs)
        used_fields = {ref.get("target_field", "") for ref in refs}
        used_dims = {ref.get("table_name", "") for ref in refs}
        dim_by_table = {dim["table_name"]: dim for dim in dimensions}
        fk_fields = [field for field in fields if field.get("role") == "fk"]

        target_dims = list(linked_dim_tables)
        if not target_dims:
            for field in fk_fields:
                matches = [dim for dim in dimensions if self._dimension_field_score(field, dim) >= 100]
                if len(matches) == 1 and matches[0]["table_name"] not in target_dims:
                    target_dims.append(matches[0]["table_name"])
                    if matrix_available:
                        self._warn(
                            f"DWD {table_name} 的 FK {field['name']} 未在总线矩阵标记，按业务键匹配到 {matches[0]['table_name']}"
                        )

        for dim_table in target_dims:
            if dim_table in used_dims:
                continue
            dim = dim_by_table.get(dim_table)
            if not dim:
                self._warn(f"DWD {table_name} 的总线矩阵引用了不存在的 DIM: {dim_table}")
                continue
            candidates = [field for field in fk_fields if field["name"] not in used_fields]
            scored = sorted(
                ((self._dimension_field_score(field, dim), field) for field in candidates),
                key=lambda item: item[0],
                reverse=True,
            )
            if not scored or scored[0][0] <= 0:
                self._warn(f"DWD {table_name} 无法为矩阵维度 {dim_table} 找到对应 FK 字段")
                continue
            score, field = scored[0]
            if len(scored) > 1 and score == scored[1][0]:
                self._warn(f"DWD {table_name} 到 {dim_table} 的 FK 匹配存在歧义，选用 {field['name']}")
            ref = {
                "entity": dim["entity"],
                "table_name": dim_table,
                "source_field": field["source_field"],
                "business_key": dim["business_key"],
                "target_field": field["name"],
            }
            refs.append(ref)
            used_fields.add(field["name"])
            used_dims.add(dim_table)

        return self._dedupe_dimension_refs(refs)

    def _dimension_field_score(self, field: Dict[str, Any], dim: Dict[str, Any]) -> int:
        field_names = {field.get("name", "").lower(), field.get("source_field", "").lower()}
        business_key = str(dim.get("business_key", "")).lower()
        business_key_source = str(dim.get("business_key_source", "")).lower()
        entity_key = f"{dim.get('entity', '').lower()}_id"
        if business_key and business_key in field_names:
            return 120
        if business_key_source and business_key_source in field_names:
            return 110
        if entity_key and entity_key in field_names:
            return 100
        entity = str(dim.get("entity", "")).lower()
        if entity and any(name.startswith(f"{entity}_") or name.endswith(f"_{entity}_id") for name in field_names):
            return 60
        return 0

    def _make_dimension_ref(
        self,
        row: Dict[str, str],
        field: Dict[str, Any],
        dim_table: str,
        dimensions: List[Dict[str, Any]],
    ) -> Dict[str, str]:
        dim = next((item for item in dimensions if item["table_name"] == dim_table), None)
        entity = dim["entity"] if dim else self._entity_from_dim_table(dim_table)
        business_key = self._pick(
            row, ["关联DIM业务键", "ref_dim_business_key", "dim_business_key"]
        ) or (dim.get("business_key") if dim else "") or field["name"]
        return {
            "entity": entity,
            "table_name": dim_table,
            "source_field": field["source_field"],
            "business_key": business_key,
            "target_field": field["name"],
        }

    def _group_join_specs(
        self,
        rows: List[Dict[str, str]],
        table_keys: List[str],
    ) -> Dict[str, List[Dict[str, str]]]:
        grouped: Dict[str, List[Dict[str, str]]] = {}
        for row in rows:
            table_name = self._pick(row, table_keys)
            source_table = self._pick(row, ["来源ODS表", "ods_table_name"])
            if not table_name or not source_table:
                continue
            grouped.setdefault(table_name, []).append({
                "source_table": source_table,
                "table_role": self._pick(row, ["表角色", "table_role"]),
                "join_type": self._pick(row, ["关联方式", "join_type"]),
                "join_condition": self._pick(row, ["关联条件", "join_condition"]),
                "remark": self._pick(row, ["备注", "remark"]),
            })
        return grouped

    def _extract_lineage_tables(self, *row_groups: List[Dict[str, str]]) -> Dict[str, Dict[str, Any]]:
        tables: Dict[str, Dict[str, Any]] = {}
        for rows in row_groups:
            for row in rows:
                table_name = self._pick(row, ["来源ODS表", "ods_table_name"]).strip()
                field_name = self._pick(row, ["来源ODS字段", "ods_column_name"]).strip()
                if not table_name or not field_name:
                    continue
                tables.setdefault(
                    table_name,
                    {"table_name": table_name, "domain": "", "description": "", "fields": []},
                )
                tables[table_name]["fields"].append(self._normalize_field({
                    "name": field_name,
                    "type": self._normalize_data_type(
                        self._pick(row, ["ODS数据类型", "ODS字段数据类型", "ods_data_type"])
                    ),
                    "description": self._pick(row, ["字段中文说明", "dwd_column_comment", "dim_column_comment"]),
                    "classification": self._pick(row, ["字段角色", "column_role"]),
                }))
        for table in tables.values():
            table["fields"] = self._dedupe_fields(table["fields"])
        return tables

    def _normalize_field(self, field: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "name": field.get("name") or field.get("field_name") or "",
            "type": field.get("type") or field.get("字段类型") or "STRING",
            "description": field.get("description") or field.get("comment") or field.get("字段说明") or "",
            "classification": field.get("classification") or field.get("字段分类") or "",
            "dimension": field.get("dimension") or field.get("维度") or "",
        }

    def _validate(self, model: Dict[str, Any]) -> None:
        if not model["dimensions"]:
            self._warn("未解析到任何 DIM 维度")
        if not model["processes"]:
            self._warn("未解析到任何 DWD 业务过程")
        dim_tables = {dim["table_name"] for dim in model["dimensions"]}
        for process in model["processes"]:
            name = process.get("business_process") or "<unknown>"
            if not process.get("grain"):
                self._warn(f"业务过程缺少粒度: {name}")
            if not process.get("grain_keys"):
                self._warn(f"业务过程缺少粒度键: {name}")
            if process.get("fact_type") != "factless" and not process.get("measures"):
                self._warn(f"业务过程缺少度量: {name}")
            for ref in process.get("dimension_refs", []):
                if ref.get("table_name") not in dim_tables:
                    self._warn(f"业务过程 {name} 引用了不存在的 DIM: {ref.get('table_name')}")
        for dim in model["dimensions"]:
            if not dim.get("business_key"):
                self._warn(f"维度缺少业务键: {dim.get('name')}")

    def _entity_from_dim_table(self, table_name: str) -> str:
        raw = str(table_name).strip()
        return raw[4:] if raw.startswith("dim_") else self._normalize_name(raw)

    def _parse_scd_type(self, value: str) -> int:
        text = str(value or "").strip().upper()
        if text in {"SCD3", "3"}:
            return 3
        if text in {"SCD2", "2"}:
            return 2
        return 1

    def _normalize_data_type(self, value: str, default_type: str = "STRING") -> str:
        raw = str(value or "").strip()
        if not raw:
            return default_type
        base = re.split(r"[\(\s]", raw.lower(), maxsplit=1)[0]
        precision_match = re.search(r"\(([^)]+)\)", raw)
        if base in {"varchar", "varchar2", "char", "nvarchar", "nvarchar2", "text", "clob"}:
            return "STRING"
        if base in {"datetime", "timestamp", "date", "time"}:
            return "STRING"
        if base == "bigint":
            return "BIGINT"
        if base in {"int", "integer"}:
            return "INT"
        if base == "smallint":
            return "SMALLINT"
        if base == "tinyint":
            return "TINYINT"
        if base in {"decimal", "numeric", "number"}:
            return f"DECIMAL({precision_match.group(1)})" if precision_match else "DECIMAL(18,2)"
        if base in {"double", "float"}:
            return base.upper()
        if base == "string":
            return "STRING"
        return raw.upper()

    def _sort_order(self, row: Dict[str, str]) -> int:
        value = self._pick(row, ["字段排序", "sort_order"])
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0

    def _unique_values(self, rows: List[Dict[str, str]], keys: List[str]) -> List[str]:
        return self._merge_unique([], [self._pick(row, keys).strip() for row in rows])

    def _merge_unique(self, first: Iterable[str], second: Iterable[str]) -> List[str]:
        result: List[str] = []
        for value in [*first, *second]:
            if value and value not in result:
                result.append(value)
        return result

    def _split_values(self, value: str) -> List[str]:
        return [item.strip() for item in re.split(r"[,，]", value or "") if item.strip()]

    def _normalize_name(self, name: str) -> str:
        raw = str(name or "").strip()
        return re.sub(r"[^a-zA-Z0-9_\u4e00-\u9fff]+", "_", raw.lower()).strip("_")

    def _dedupe_fields(self, fields: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        seen = set()
        result = []
        for field in fields:
            name = field.get("name", "")
            if not name or name in seen:
                continue
            seen.add(name)
            result.append(field)
        return result

    def _dedupe_dimension_refs(self, refs: List[Dict[str, str]]) -> List[Dict[str, str]]:
        seen = set()
        result = []
        for ref in refs:
            key = (ref.get("table_name"), ref.get("target_field"))
            if key in seen:
                continue
            seen.add(key)
            result.append(ref)
        return result

    def _is_positive_marker(self, value: str) -> bool:
        return str(value or "").strip().upper() in {"Y", "YES", "TRUE", "1", "✓", "√"}

    def _pick(self, row: Dict[str, str], keys: List[str]) -> str:
        for key in keys:
            if key in row and row[key]:
                return self._stringify(row[key])
        return ""

    def _clean_row(self, row: Dict[str, Any]) -> Dict[str, str]:
        return {self._stringify(key): self._stringify(value) for key, value in row.items() if key}

    def _stringify(self, value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _warn(self, message: str) -> None:
        if message in self.warnings:
            return
        self.warnings.append(message)
        if self.logger:
            self.logger.warning(message)
