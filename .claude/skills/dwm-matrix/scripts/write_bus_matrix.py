#!/usr/bin/env python3
"""
Bus matrix Excel generator for dwm-matrix skill.

Reads upstream CSV outputs and generates a cross-tab xlsx file.
Bus matrix grid (fact × dimension) is derived from field metadata FK relations + dim_spec,
no separate fact-dim-ref CSV needed.

Usage:
    python write_bus_matrix.py \
        --business-process output/dwm-bus-matrix/dwm_bp_business_process.csv \
        --subject-area  output/dwm-bus-matrix/dwm_bp_subject_area.csv \
        --dim-registry  output/dwm-bus-matrix/dwm_dim_spec.csv \
        --field-metadata output/metadata_parse/all_tables_metadata.xlsx \
        --output        output/dwm-bus-matrix/dwm_bus_matrix.xlsx \
        --version       v1.0

Requires: openpyxl (pip install openpyxl)
"""

import sys
import os
import argparse
from pathlib import Path

# Reuse shared read_csv/read_xlsx to ensure consistent handling
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../dwm-shared/scripts"))
from read_csv import read_csv
from read_xlsx import read_xlsx

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def build_matrix(
    table_profile: list[dict],
    subject_area: list[dict],
    dim_spec: list[dict],
    field_profile: list[dict],
    dwd_fact_spec: list[dict] | None = None,
) -> tuple[list[dict], list[dict], dict[tuple[str, str], str]]:
    """
    Build bus matrix data structures.
    Grid cells are derived from:
      1. field_profile FK relations (if 字段角色='foreign_key' exists)
      2. dwd_fact_spec FK fields matched to dim_spec PK fields (fallback)

    Returns:
        - rows: list of fact business processes (sorted by subject area)
        - columns: list of unique conformed dimensions (deduplicated from field-level dim_spec)
        - cells: dict of (bp_standard_name, dim_table_name) -> marker
    """
    # Build subject area lookup
    sa_lookup = {sa["主题域编码"]: sa for sa in subject_area}

    # Build rows: one row per business process
    rows = []
    seen_bp = {}
    ods_to_bp = {}
    for tp in table_profile:
        bp_name = tp.get("业务过程英文名称", "")
        ods_table = tp.get("涉及ODS表", "")
        if ods_table:
            ods_to_bp[ods_table] = bp_name
        if bp_name in seen_bp:
            continue
        sa_code = tp.get("主题域编码", "")
        sa_name = sa_lookup.get(sa_code, {}).get("中文名称", sa_code)
        seen_bp[bp_name] = True
        rows.append({
            "subject_area_code": sa_code,
            "subject_area_name": sa_name,
            "business_process": tp.get("业务过程中文名称", ""),
            "bp_standard_name": bp_name,
            "dwd_table_name": f"dwd_{sa_code.lower()}_{bp_name}_df",
            "fact_type": tp.get("事实表类型", ""),
            "grain_statement": tp.get("粒度声明", ""),
        })

    rows.sort(key=lambda r: (r["subject_area_code"], r["bp_standard_name"]))

    # Build columns: deduplicate dim_spec (field-level) to unique dimensions
    seen_dims = {}
    for row in dim_spec:
        dim_key = row.get("维度表名", "")
        if dim_key and dim_key not in seen_dims:
            seen_dims[dim_key] = {
                "维度表名": dim_key,
                "维度中文名称": row.get("维度中文名称", ""),
                "来源ODS表": row.get("来源ODS表", ""),
            }
    columns = sorted(seen_dims.values(), key=lambda d: d["维度表名"])

    # Build ODS source table -> dim_table_name mapping
    src_to_dim = {}
    for dim in columns:
        src_table = dim.get("来源ODS表", "")
        if src_table:
            src_to_dim[src_table] = dim["维度表名"]

    # Build cells: derive from field metadata FK relations
    cells: dict[tuple[str, str], str] = {}
    for fp in field_profile:
        if fp.get("字段角色", "") != "foreign_key":
            continue
        ods_table = fp.get("表名", "")
        bp_name = ods_to_bp.get(ods_table, "")
        if not bp_name:
            continue
        ref_info = fp.get("外键引用", "")
        if not ref_info:
            continue
        ref_table_short = ref_info.split(".")[0].split("(")[0].strip()
        for src_table, dim_key in src_to_dim.items():
            if src_table.endswith(ref_table_short):
                cells[(bp_name, dim_key)] = "Y"
                break

    # Fallback: derive from dwd_fact_spec FK fields matched to dim PK fields
    if not cells and dwd_fact_spec:
        # Build dim PK field -> dim_table_name mapping
        pk_to_dim: dict[str, str] = {}
        for row in dim_spec:
            if row.get("字段角色", "") == "pk":
                pk_to_dim[row.get("字段名", "")] = row.get("维度表名", "")

        for row in dwd_fact_spec:
            if row.get("字段角色", "") != "fk":
                continue
            bp_name = row.get("业务过程标准名", "")
            fk_field = row.get("字段名", "")
            dim_key = pk_to_dim.get(fk_field, "")
            if bp_name and dim_key:
                cells[(bp_name, dim_key)] = "Y"

    return rows, columns, cells


def write_xlsx(
    rows: list[dict],
    columns: list[dict],
    cells: dict[tuple[str, str], str],
    output_path: str,
    version: str = "v1.0",
):
    """Generate the bus matrix Excel file."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"总线矩阵 {version}"

    # -- Styles --
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    check_font = Font(size=12, bold=True, color="2E7D32")
    dash_font = Font(size=12, color="BDBDBD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # -- Header row --
    fixed_headers = ["主题域", "业务过程", "业务过程代码", "粒度声明", "事实表名称", "事实表类型"]
    dim_headers = [d["维度中文名称"] or d["维度表名"] for d in columns]
    all_headers = fixed_headers + dim_headers

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # -- Data rows --
    data_start_row = 2

    for row_idx, row_data in enumerate(rows, data_start_row):
        sa_code = row_data["subject_area_code"]
        sa_name = row_data["subject_area_name"]
        bp_cn = row_data["business_process"]
        bp_name = row_data["bp_standard_name"]
        dwd_name = row_data["dwd_table_name"]
        grain = row_data["grain_statement"]
        fact_type = row_data["fact_type"]

        ws.cell(row=row_idx, column=1, value=f"{sa_name}({sa_code})").border = thin_border
        ws.cell(row=row_idx, column=2, value=bp_cn).border = thin_border
        ws.cell(row=row_idx, column=3, value=bp_name).border = thin_border
        cell_d = ws.cell(row=row_idx, column=4, value=grain)
        cell_d.alignment = Alignment(wrap_text=True, vertical="center")
        cell_d.border = thin_border
        ws.cell(row=row_idx, column=5, value=dwd_name).border = thin_border
        ws.cell(row=row_idx, column=6, value=fact_type).border = thin_border

        for dim_idx, dim in enumerate(columns):
            dim_key = dim["维度表名"]
            marker = cells.get((bp_name, dim_key), "N")
            cell = ws.cell(row=row_idx, column=7 + dim_idx, value=marker)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = check_font if marker == "Y" else dash_font

    # -- Column widths --
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18
    for i in range(len(columns)):
        col_letter = get_column_letter(7 + i)
        ws.column_dimensions[col_letter].width = 14

    wb.save(output_path)
    print(f"Bus matrix written -> {output_path} ({len(rows)} processes × {len(columns)} dimensions)")


def main():
    parser = argparse.ArgumentParser(description="Generate bus matrix Excel")
    parser.add_argument("--business-process", required=True, help="Path to dwm_bp_business_process.csv")
    parser.add_argument("--subject-area", required=True, help="Path to dwm_bp_subject_area.csv")
    parser.add_argument("--dim-registry", required=True, help="Path to dwm_dim_spec.csv")
    parser.add_argument("--field-metadata", required=True, help="Path to all_tables_metadata.xlsx")
    parser.add_argument("--dwd-fact-spec", default=None,
                        help="Path to dwm_dwd_fact_spec.csv (fallback for FK relation derivation)")
    parser.add_argument("--output", "-o", default="output/dwm-bus-matrix/dwm_bus_matrix.xlsx",
                        help="Output xlsx path (default: output/dwm-bus-matrix/dwm_bus_matrix.xlsx)")
    parser.add_argument("--version", "-v", default="v1.0", help="Matrix version (default: v1.0)")
    args = parser.parse_args()

    table_profile = read_csv(args.business_process)
    subject_area = read_csv(args.subject_area)
    dim_spec = read_csv(args.dim_registry)
    field_profile = read_xlsx(args.field_metadata)
    dwd_fact_spec = read_csv(args.dwd_fact_spec) if args.dwd_fact_spec else None

    rows, columns, cells = build_matrix(table_profile, subject_area, dim_spec, field_profile, dwd_fact_spec)
    write_xlsx(rows, columns, cells, args.output, args.version)


if __name__ == "__main__":
    main()
